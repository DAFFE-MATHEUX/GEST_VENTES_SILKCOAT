from datetime import datetime
from django.template import TemplateDoesNotExist
from django.contrib import messages
from django.shortcuts import redirect, render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from gest_entreprise.models import Entreprise
from django.utils.timezone import now
from decimal import Decimal
import qrcode
from io import BytesIO
import base64
import openpyxl
from openpyxl.utils import get_column_letter

from gestion_notifications.models import Notification
from .utils import *
from gestion_audit.views import enregistrer_audit
from .models import * 
from django.core.mail import EmailMessage
from django.utils import timezone
from django.conf import settings
from django.http import HttpResponse
from django.db.models import Sum, F, Count, Q, ExpressionWrapper, IntegerField
from openpyxl import Workbook

from django.db import transaction
from collections import defaultdict
from django.db.models import Sum, Count, F
from urllib.parse import urlencode

from django.db import IntegrityError, DatabaseError
from django.views.decorators.csrf import csrf_protect
from django.contrib.auth.decorators import login_required

import logging
logger = logging.getLogger(__name__)

from django.db.models.signals import post_save, post_delete
from django.dispatch import receiver
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers


#================================================================================================
# Fonction pour ajouter une catégorie de produit
#================================================================================================

@login_required
@csrf_protect
def ajouter_categorie(request):
    try:
        if request.method == 'POST':
            nom = request.POST.get('nom', '').strip()
            description = request.POST.get('description', '').strip()

            # 1️⃣ Validation du champ obligatoire
            if not nom:
                messages.error(request, "Le nom de la catégorie est obligatoire.")
                return redirect('produits:ajouter_categorie')

            # 2️⃣ Vérification de doublon
            if CategorieProduit.objects.filter(desgcategorie__iexact=nom).exists():
                messages.warning(
                    request,
                    "Cette catégorie existe déjà."
                )
                return redirect('produits:ajouter_categorie')

            # 3️⃣ Création sécurisée
            CategorieProduit.objects.create(
                desgcategorie=nom,
                description=description
            )

            messages.success(request, "Catégorie ajoutée avec succès.")
            return redirect('produits:listes_categorie')

        # 4️⃣ Mauvaise méthode HTTP
        messages.error(request, "Méthode non autorisée.")
        return redirect('produits:listes_categorie')

    except IntegrityError:
        messages.error(
            request,
            "Erreur d'intégrité : données invalides ou doublon détecté."
        )
        return redirect('produits:ajouter_categorie')

    except DatabaseError:
        messages.error(
            request,
            "Erreur de base de données. Veuillez réessayer plus tard."
        )
        return redirect('produits:ajouter_categorie')

    except Exception as e:
        # 5️⃣ Erreur inconnue (loggable)
        messages.error(
            request,
            "Une erreur inattendue est survenue."
        )
        return redirect('produits:ajouter_categorie')

#================================================================================================
# Fonction pour éffectuer une nouvelle vente
#================================================================================================

def generer_code_vente():
    prefix = "VENTE"
    date_str = timezone.now().strftime('%Y%m%d')

    last_vente = VenteProduit.objects.order_by('-id').first()

    if last_vente:
        try:
            dernier_numero = int(last_vente.code.split('-')[-1])
        except (IndexError, ValueError):
            dernier_numero = 0
    else:
        dernier_numero = 0

    nouveau_numero = dernier_numero + 1

    return f"{prefix}-{date_str}-{str(nouveau_numero).zfill(4)}"


@login_required
@csrf_protect
def vendre_produit(request):
    try:
        produits = Produits.objects.all().order_by("desgprod")

        if request.method == "POST":

            ids = request.POST.getlist("produit_id[]")
            quantites = request.POST.getlist("quantite[]")
            reductions = request.POST.getlist("reduction[]")

            nom_complet = request.POST.get("nom_complet_client", "").strip()
            telephone = request.POST.get("telephone_client", "").strip()
            adresse = request.POST.get("adresse_client", "").strip()

            if not nom_complet or not telephone or not adresse:
                messages.error(
                    request,
                    "Veuillez renseigner le nom, le téléphone et l'adresse du client."
                )
                return redirect("produits:vendre_produit")

            total_general = 0
            lignes = []
            produits_sans_stock = []

            with transaction.atomic():

                # 1️⃣ Création de la vente
                vente = VenteProduit.objects.create(
                    code = generer_code_vente(),
                    total=0,
                    benefice_total=0,
                    utilisateur=request.user,
                    nom_complet_client=nom_complet,
                    telclt_client=telephone,
                    adresseclt_client=adresse
                )

                # 2️⃣ Validation des produits
                for prod_id, qte_str, red_str in zip(ids, quantites, reductions):

                    if not prod_id:
                        continue

                    try:
                        produit = Produits.objects.get(id=int(prod_id))
                    except (Produits.DoesNotExist, ValueError):
                        continue

                    try:
                        quantite = int(qte_str or 0)
                        reduction = int(red_str or 0)
                    except ValueError:
                        raise IntegrityError("Quantité ou réduction invalide")

                    if quantite <= 0:
                        continue

                    stock = StockProduit.objects.select_for_update().filter(
                        produit=produit
                    ).first()

                    if not stock or stock.qtestock <= 0:
                        produits_sans_stock.append(produit.desgprod)
                        continue

                    if stock.qtestock < quantite:
                        messages.error(
                            request,
                            f"Stock insuffisant pour {produit.desgprod} "
                            f"(Disponible : {stock.qtestock})"
                        )
                        raise IntegrityError()

                    if reduction > produit.pu:
                        reduction = produit.pu

                    prix_net = produit.pu - reduction
                    sous_total = prix_net * quantite
                    total_general += sous_total

                    lignes.append({
                        "produit": produit,
                        "quantite": quantite,
                        "pu": produit.pu,
                        "reduction": reduction,
                        "stock": stock
                    })

                if not lignes:
                    messages.error(request, "Aucun produit valide sélectionné.")
                    vente.delete()
                    return redirect("produits:vendre_produit")

                # 3️⃣ Création lignes + stock + détails email
                details_lignes = []
                total_qte = 0
                total_reduction = 0

                for ligne in lignes:

                    LigneVente.objects.create(
                        vente=vente,
                        produit=ligne["produit"],
                        quantite=ligne["quantite"],
                        prix=ligne["pu"],
                        montant_reduction=ligne["reduction"],
                    )

                    stock = ligne["stock"]
                    stock.qtestock -= ligne["quantite"]
                    stock.save(update_fields=["qtestock"])

                    # 🔔 Alerte stock
                    if stock.qtestock <= stock.seuil:
                        Notification.objects.create(
                            destinataire=None,
                            destinataire_email=settings.ADMIN_EMAIL,
                            titre="⚠️ Stock critique",
                            message=(
                                f"Le produit '{ligne['produit'].desgprod}' "
                                f"est presque en rupture.\n"
                                f"Stock restant : {stock.qtestock}"
                            )
                        )

                    benefice_ligne = (
                        (ligne["pu"] - ligne["produit"].prix_en_gros - ligne["reduction"])
                        * ligne["quantite"]
                    )

                    sous_total = (ligne["pu"] - ligne["reduction"]) * ligne["quantite"]

                    details_lignes.append(
                        f"Catégorie : {ligne['produit'].categorie.desgcategorie}\n"
                        f"Produit : {ligne['produit'].desgprod}\n"
                        f"Quantité : {ligne['quantite']}\n"
                        f"PU : {ligne['pu']}\n"
                        f"Réduction unitaire : {ligne['reduction']}\n"
                        f"Sous-total : {sous_total}\n"
                        f"Bénéfice ligne : {benefice_ligne}\n"
                        f"---------------------------------\n"
                    )

                    total_qte += ligne["quantite"]
                    total_reduction += ligne["reduction"]

                # 4️⃣ Mise à jour totaux
                vente.total = total_general
                vente.calculer_totaux()
                vente.save(update_fields=["total", "benefice_total"])

                # 5️⃣ Notification + email
                Notification.objects.create(
                    destinataire=None,
                    destinataire_email=settings.ADMIN_EMAIL,
                    titre=f"Nouvelle vente {vente.code}",
                    message=(
                        f"Vente réalisée par {request.user.get_full_name() or "Admin"} "
                        f"pour {nom_complet} | Montant : {vente.total} GNF"
                    )
                )

                email_body = (
                    f"Nouvelle vente : {vente.code}\n\n"
                    f"Client : {vente.nom_complet_client}\n"
                    f"Téléphone : {vente.telclt_client}\n"
                    f"Adresse : {vente.adresseclt_client}\n\n"
                    f"-------- DÉTAILS DE LA VENTE --------\n"
                    f"{''.join(details_lignes)}\n"
                    f"------------------------------------\n"
                    f"TOTAL QUANTITÉ : {total_qte}\n"
                    f"TOTAL RÉDUCTION : {total_reduction}\n"
                    f"TOTAL MONTANT : {vente.total}\n"
                    f"BÉNÉFICE GLOBAL : {vente.benefice_total}\n"
                    f"VENTE RÉALISÉE PAR : {request.user.get_full_name() or 'Admin'}\n"
                )

                EmailMessage(
                    subject=f"Nouvelle vente : {vente.code}",
                    body=email_body,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[settings.ADMIN_EMAIL]
                ).send(fail_silently=False)

            messages.success(request, "✅ Vente enregistrée avec succès.")
            return redirect(
                reverse(
                    "produits:recu_vente_global",
                    kwargs={"vente_code": vente.code}
                )
            )

        return render(
            request,
            "gestion_produits/ventes/nouvelle_vente.html",
            {"produits": produits}
        )

    except IntegrityError as ie:
        messages.error(request, f"Erreur d'intégrité : {str(ie)}")
        return redirect("produits:vendre_produit")

    except DatabaseError as de:
        messages.error(request, f"Erreur base de données : {str(de)}")
        return redirect("produits:vendre_produit")

    except Exception as e:
        logger.exception("Erreur vente produit")
        messages.error(request, f"Erreur inattendue : {str(e)}")
        return redirect("produits:vendre_produit")

#==========================================================
# Fonction pour le retour de vente
#==========================================================
@login_required
@csrf_protect
def enregistrer_retour(request, ligne_id):
    ligne = get_object_or_404(LigneVente, id=ligne_id)
    vente = ligne.vente
    quantite_restante = ligne.quantite_restante

    if quantite_restante <= 0:
        messages.warning(
            request,
            "Toutes les unités de ce produit ont déjà été retournées."
        )
        return redirect('produits:listes_des_ventes')

    if request.method == 'POST':
        try:
            quantite_retour = int(request.POST.get('quantite_retour', 0))
            motif = request.POST.get('motif', '').strip()

            if quantite_retour < 1:
                messages.error(request, "La quantité à retourner doit être au moins 1.")
                return redirect(request.path)

            if quantite_retour > quantite_restante:
                messages.error(
                    request,
                    f"Vous ne pouvez retourner que {quantite_restante} unité(s)."
                )
                return redirect(request.path)

            # 1️⃣ Création du retour
            retour = RetourVente.objects.create(
                ligne_vente=ligne,
                quantite_retour=quantite_retour,
                motif=motif,
                date_retour=timezone.now()
            )

            # 2️⃣ Recalcul des totaux
            vente.calculer_totaux()
            vente.save(update_fields=["total", "benefice_total"])

            # 3️⃣ Calculs retour
            prix_net = ligne.prix - ligne.montant_reduction
            montant_retour = prix_net * quantite_retour
            
            # 🔹 Mise à jour stock
            stock, created = StockProduit.objects.get_or_create(
                produit = ligne.produit,
                defaults = {"qtestock": quantite_retour}
            )
            if not created:
                stock.qtestock += quantite_retour
                stock.save(update_fields=["qtestock"])

            # 4️⃣ EMAIL RETOUR
            try:
                email_body = (
                    f"RETOUR DE PRODUIT – {vente.code}\n\n"
                    f"Client : {vente.nom_complet_client}\n"
                    f"Téléphone : {vente.telclt_client}\n"
                    f"Adresse : {vente.adresseclt_client}\n\n"
                    f"PRODUIT RETOURNÉ\n"
                    f"-----------------------------------\n"
                    f"- Produit : {ligne.produit.desgprod}\n"
                    f"  Quantité retournée : {quantite_retour}\n"
                    f"  Prix unitaire : {ligne.prix}\n"
                    f"  Réduction unitaire : {ligne.montant_reduction}\n"
                    f"  Montant retourné : {montant_retour}\n"
                    f"  Motif : {motif or 'Non précisé'}\n\n"
                    f"-----------------------------------\n"
                    f"TOTAL RETOUR : {montant_retour} GNF\n"
                    f"DATE DU RETOUR : {timezone.now().strftime('%d/%m/%Y %H:%M')}\n"
                )

                EmailMessage(
                    subject=f"Retour produit – {vente.code}",
                    body=email_body,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[settings.ADMIN_EMAIL]
                ).send(fail_silently=False)
                
                #  Notification du Retour de Vente
                Notification.objects.create(
                    destinataire=None,
                    destinataire_email = settings.ADMIN_EMAIL,
                    titre=f"Retour de vente {vente.code}",
                    message=(
                        f"Vente retourer du Client : {vente.nom_complet_client} \n"
                        f"Quantité retournée : {quantite_retour}\n",
                        f"Montant : {montant_retour} GNF \n",
                        f"DATE DU RETOUR : {timezone.now().strftime('%d/%m/%Y %H:%M')}\n"
                    )
                )

            except Exception as e:
                logger.warning(f"Email retour non envoyé : {e}")

            messages.success(
                request,
                f"{quantite_retour} unité(s) retournée(s) avec succès."
            )
            return redirect('produits:listes_des_ventes')

        except ValueError:
            messages.error(request, "Quantité invalide.")

    return render(
        request,
        'gestion_produits/ventes/enregistrer_retour.html',
        {
            'ligne': ligne,
            'quantite_restante': quantite_restante,
            'today': timezone.now()
        }
    )

#================================================================================================
# Fonction pour afficher l'historique des ventes par date
#================================================================================================

@login_required
@csrf_protect
def historique_ventes(request):
    ventes = (
        VenteProduit.objects
        .select_related("utilisateur")
        .prefetch_related("lignes__produit__categorie")
        .order_by("-date_vente")
    )

    ventes_par_date = defaultdict(list)

    for vente in ventes:
        ventes_par_date[vente.date_vente.date()].append(vente)

    historique = []

    for date, ventes_du_jour in ventes_par_date.items():
        total_montant = 0
        total_quantite = 0
        total_benefice = 0
        categories = set()

        for vente in ventes_du_jour:
            for ligne in vente.lignes.all():

                total_quantite += ligne.quantite
                total_benefice += ligne.benefice

                if ligne.produit and ligne.produit.categorie:
                    categories.add(ligne.produit.categorie.id)

            total_montant += vente.total

        historique.append({
            "date": date,
            "ventes": ventes_du_jour,
            "total_montant": total_montant,
            "total_quantite": total_quantite,
            "total_categories": len(categories),
            "total_profit": total_benefice,
        })

    return render(
        request,
        "gestion_produits/ventes/historique_ventes.html",
        {"historique": historique}
    )

#================================================================================================
# Fonction pour afficher l'historique des commandes et livraisons par date
#================================================================================================
@csrf_protect
@login_required
def historique_commandes_livraisons(request):
    """
    Vue sécurisée pour afficher l'historique des commandes et livraisons
    avec calculs côté Python.
    """
    try:
        historique = []
        commandes = Commandes.objects.all().order_by('-datecmd')

        total_commandes = 0
        total_livrees = 0
        total_restantes = 0

        for cmd in commandes:
            # 🔹 Récupérer les livraisons liées à cette commande
            livraisons = LivraisonsProduits.objects.filter(commande=cmd).order_by('datelivrer')
            
            # 🔹 Total livré pour cette commande
            total_livree = livraisons.aggregate(total=Sum('qtelivrer'))['total'] or 0
            
            # 🔹 Quantité restante
            qte_restante = max(cmd.qtecmd - total_livree, 0)

            historique.append({
                'commande': cmd,
                'livraisons': livraisons,
                'total_livree': total_livree,
                'qte_restante': qte_restante
            })

            # 🔹 Totaux pour le footer
            total_commandes += cmd.qtecmd or 0
            total_livrees += total_livree
            total_restantes += qte_restante

        context = {
            'historique': historique,
            'total_commandes': total_commandes,
            'total_livrees': total_livrees,
            'total_restantes': total_restantes
        }

        return render(
            request,
            'gestion_produits/livraisons/historique_commandes_livraisons.html',
            context
        )

    except DatabaseError as db_err:
        logger.error(f"Erreur base de données historique commandes/livraisons: {db_err}")
        messages.error(request, "Erreur lors de la récupération des commandes/livraisons.")
        return render(request, 'gestion_produits/livraisons/historique_commandes_livraisons.html', {
            'historique': [],
            'total_commandes': 0,
            'total_livrees': 0,
            'total_restantes': 0
        })

    except Exception as e:
        logger.exception("Erreur inattendue historique commandes/livraisons")
        messages.error(request, "Une erreur inattendue est survenue.")
        return render(request, 'gestion_produits/livraisons/historique_commandes_livraisons.html', {
            'historique': [],
            'total_commandes': 0,
            'total_livrees': 0,
            'total_restantes': 0
        })

#================================================================================================
# Fonction pour éffectuer une nouvelle commande
#================================================================================================
def generer_code_commandes():
    prefix = "CMD"
    date_str = timezone.now().strftime('%Y%m%d')

    last_commande = Commandes.objects.order_by('-id').first()

    if last_commande:
        try:
            dernier_numero = int(last_commande.numcmd.split('-')[-1])
        except (IndexError, ValueError):
            dernier_numero = 0
    else:
        dernier_numero = 0

    nouveau_numero = dernier_numero + 1

    return f"{prefix}-{date_str}-{str(nouveau_numero).zfill(4)}"


@csrf_protect
@login_required

def nouvelle_commande(request):
    """
    Création sécurisée d'une nouvelle commande fournisseur
    + envoi email détaillé à l'admin
    """

    produits = Produits.objects.all()
    produits_data = [{"produit": p} for p in produits]

    stock_total = StockProduit.objects.aggregate(
        total=Sum('qtestock')
    )['total'] or 0

    if request.method == "POST":
        ids = request.POST.getlist("produit_id[]")
        quantites = request.POST.getlist("quantite[]")

        nom_complet_fournisseur = request.POST.get("nom_complet_fournisseur", "").strip()
        telephone_fournisseur = request.POST.get("telephone_fournisseur", "").strip()
        adresse_fournisseur = request.POST.get("adresse_fournisseur", "").strip()

        if not nom_complet_fournisseur or not telephone_fournisseur or not adresse_fournisseur:
            messages.error(request, "Veuillez renseigner toutes les informations du fournisseur.")
            return redirect("produits:nouvelle_commande")

        if not ids or not quantites:
            messages.error(request, "Aucun produit sélectionné.")
            return redirect("produits:nouvelle_commande")

        lignes = []
        total_general = 0
        numero_commande = generer_code_commandes()

        try:
            with transaction.atomic():

                for prod_id, qte_str in zip(ids, quantites):
                    try:
                        produit = Produits.objects.get(id=prod_id)
                    except Produits.DoesNotExist:
                        continue

                    try:
                        qte = int(qte_str or 0)
                    except ValueError:
                        continue

                    if qte <= 0:
                        continue

                    # Enregistrement de la ligne de commande
                    Commandes.objects.create(
                        numcmd=numero_commande,
                        qtecmd=qte,
                        produits=produit,
                        nom_complet_fournisseur=nom_complet_fournisseur,
                        adresse_fournisseur=adresse_fournisseur,
                        telephone_fournisseur=telephone_fournisseur,
                    )

                    sous_total = produit.pu * qte
                    total_general += sous_total

                    lignes.append({
                        "produit": produit.desgprod,
                        "qte": qte,
                        "pu": produit.pu,
                        "sous_total": sous_total
                    })

                if not lignes:
                    messages.error(request, "Aucune ligne de commande valide.")
                    return redirect("produits:nouvelle_commande")

                # ================= EMAIL =================
                try:
                    email_body = (
                        f"📦 NOUVELLE COMMANDE FOURNISSEUR\n\n"
                        f"Numéro de commande : {numero_commande}\n\n"
                        f"FOURNISSEUR\n"
                        f"Nom : {nom_complet_fournisseur}\n"
                        f"Téléphone : {telephone_fournisseur}\n"
                        f"Adresse : {adresse_fournisseur}\n\n"
                        f"---------------- DÉTAILS ----------------\n"
                    )

                    for l in lignes:
                        email_body += (
                            f"- Produit : {l['produit']}\n"
                            f"  Quantité : {l['qte']}\n"
                            f"  PU : {l['pu']} GNF\n"
                            f"  Sous-total : {l['sous_total']} GNF\n\n"
                        )

                    email_body += (
                        f"----------------------------------------\n"
                        f"TOTAL GÉNÉRAL : {total_general} GNF\n"
                    )
                    # 🔹 Notification Django
                    Notification.objects.create(
                        destinataire=request.user,  # facultatif, ou None si notification globale
                        destinataire_email=settings.ADMIN_EMAIL,
                        titre=f"Nouvelle Commande {numero_commande}",
                        message=(
                            f"Commande réalisée par {request.user.get_full_name() or 'Admin'} "
                            f"avec le fournisseur {nom_complet_fournisseur} | Quantite : {l['qte']} GNF"
                        )
                    )
                    EmailMessage(
                        subject=f"📦 Nouvelle commande {numero_commande}",
                        body=email_body,
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        to=[settings.ADMIN_EMAIL],
                    ).send(fail_silently=False)

                except Exception as e:
                    logger.warning(f"Email non envoyé pour la commande {numero_commande}: {e}")
                    messages.warning(
                        request,
                        "Commande enregistrée mais l'email n'a pas été envoyé."
                    )

            messages.success(
                request,
                f"✅ Commande {numero_commande} enregistrée avec succès."
            )
            return redirect("produits:listes_des_commandes")

        except Exception as e:
            logger.exception("Erreur lors de la création de la commande")
            messages.error(
                request,
                "❌ Une erreur est survenue lors de l'enregistrement de la commande."
            )
            return redirect("produits:nouvelle_commande")

    return render(
        request,
        "gestion_produits/commandes/nouvelle_commande.html",
        {
            "produits_data": produits_data,
            "stock_total": stock_total,
        }
    )

#================================================================================================
# Fonction pour éffectuer une receptin de livraisons des commandes
#================================================================================================
def generer_code_livraisons():
    prefix = "LIV"
    date_str = timezone.now().strftime('%Y%m%d')

    last_livraison = LivraisonsProduits.objects.order_by('-id').first()

    if last_livraison:
        try:
            dernier_numero = int(last_livraison.numlivrer.split('-')[-1])
        except (IndexError, ValueError):
            dernier_numero = 0
    else:
        dernier_numero = 0

    nouveau_numero = dernier_numero + 1

    return f"{prefix}-{date_str}-{str(nouveau_numero).zfill(4)}"

@login_required
@transaction.atomic
def reception_livraison(request):
    """
    Réception des commandes NON encore totalement livrées
    Avec possibilité d'annuler une partie de la commande
    """
    commandes_data = []

    # 🔹 Récupérer uniquement les commandes non totalement livrées
    commandes = Commandes.objects.exclude(statuts="Livrée").order_by("-datecmd")

    for cmd in commandes:
        total_livree = (
            LivraisonsProduits.objects
            .filter(commande=cmd)
            .aggregate(total=Sum("qtelivrer"))["total"] or 0
        )
        qte_restante = cmd.qtecmd - total_livree

        if qte_restante > 0:
            commandes_data.append({
                "commande": cmd,
                "total_livree": total_livree,
                "qte_restante": qte_restante
            })

    # 🔹 Traitement POST
    if request.method == "POST":
        commande_ids = request.POST.getlist("commande_id[]")
        qte_livree_list = request.POST.getlist("qte_livree[]")

        if len(commande_ids) != len(qte_livree_list):
            messages.error(request, "Erreur : données invalides.")
            return redirect("produits:reception_livraison")

        livraisons_effectuees = []

        for i, cmd_id in enumerate(commande_ids):
            try:
                cmd = Commandes.objects.get(id=cmd_id)
                qte_livree = int(qte_livree_list[i])
                # 🔹 Récupérer la quantité à annuler depuis POST
                qte_annuler = int(request.POST.get(f"qte_annuler_{cmd.id}", 0))
            except (Commandes.DoesNotExist, ValueError):
                continue

            # 🔹 Calcul de la quantité restante
            total_livree = (
                LivraisonsProduits.objects
                .filter(commande=cmd)
                .aggregate(total=Sum("qtelivrer"))["total"] or 0
            )
            qte_restante = cmd.qtecmd - total_livree

            # 🔹 Sécurité : ne pas dépasser le restant
            if qte_livree < 0 or qte_livree > qte_restante:
                qte_livree = min(max(qte_livree, 0), qte_restante)
            if qte_annuler < 0 or qte_annuler > qte_restante:
                qte_annuler = min(max(qte_annuler, 0), qte_restante)

            if qte_livree == 0 and qte_annuler == 0:
                continue  # rien à faire pour cette commande

            # 🔹 Enregistrement livraison
            if qte_livree > 0:
                LivraisonsProduits.objects.create(
                    numlivrer = generer_code_livraisons(),
                    commande=cmd,
                    produits=cmd.produits,
                    qtelivrer=qte_livree,
                    datelivrer=timezone.now().date(),
                    statuts="Livrée"
                )

                # 🔹 Mise à jour stock
                stock, created = StockProduit.objects.get_or_create(
                    produit=cmd.produits,
                    defaults={"qtestock": qte_livree}
                )
                if not created:
                    stock.qtestock += qte_livree
                    stock.save(update_fields=["qtestock"])

            # 🔹 Mise à jour commande : retirer quantité annulée
            cmd.qtecmd -= qte_annuler
            total_livree += qte_livree
            if total_livree >= cmd.qtecmd:
                cmd.statuts = "Livrée"
            elif total_livree > 0:
                cmd.statuts = "Partiellement livrée"
            else:
                cmd.statuts = "Non Livrée"
            cmd.save(update_fields=["qtecmd", "statuts"])

            # 🔹 Historique pour notification
            livraisons_effectuees.append({
                "commande": cmd.numcmd,
                "produit": cmd.produits.desgprod,
                "qte_livree": qte_livree,
                "qte_annuler": qte_annuler,
                "fournisseur": cmd.nom_complet_fournisseur
            })

        # 🔹 Notification email
        if livraisons_effectuees:
            contenu = "📦 Nouvelle réception de livraison :\n\n"
            for l in livraisons_effectuees:
                contenu += (
                    f"- Commande : {l['commande']} | "
                    f"Produit : {l['produit']} | "
                    f"Livrée : {l['qte_livree']} | "
                    f"Annulée : {l['qte_annuler']} | "
                    f"Fournisseur : {l['fournisseur']}\n"
                )

            EmailMessage(
                subject="Nouvelle réception de livraison",
                body=contenu,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[settings.ADMIN_EMAIL]
            ).send(fail_silently=True)
        # Notification
        try:
            Notification.objects.create(
                destinataire=None,  # ou request.user si tu veux lier la notif à l'utilisateur
                destinataire_email=settings.ADMIN_EMAIL,
                titre=f"Nouvelle Livraison {l['commande']}",
                message=(
                    f"Livraison enregistrée par {request.user.get_full_name() or 'Admin'} | "
                    f"Commande {l['commande']} | Produit : {l['produit']} | "
                    f"Livrée : {l['qte_livree']} | Annulée : {l['qte_annuler']}"
                )
            )
        except Exception as e:
            logger.warning(f"Notification non créée pour livraison {l['commande']}: {e}")
        messages.success(request, "Livraison enregistrée avec succès.")
        return redirect("produits:listes_des_livraisons")

    return render(
        request,
        "gestion_produits/livraisons/reception_livraison.html",
        {"commandes": commandes_data}
    )

#=============================================================================================
# Fonction pour gérer les réçu Global de Ventes
#=============================================================================================
@login_required
def recu_vente_global(request, vente_code):
    try:
        vente = VenteProduit.objects.get(code=vente_code)
    except VenteProduit.DoesNotExist:
        messages.error(request, f"Aucune vente ne correspond au code : {vente_code}")
        return redirect("produits:listes_des_ventes")
    except Exception as ex:
        messages.error(request, f"Erreur inattendue : {str(ex)}")
        return redirect("produits:listes_des_ventes")

    # --- récupérer les lignes ---
    lignes = LigneVente.objects.filter(vente=vente)
    if not lignes.exists():
        messages.error(request, "Aucun produit trouvé pour cette vente.")
        return redirect("produits:listes_des_ventes")
    
    has_retour = lignes.filter(retours__isnull=False).exists()
    total_quantite_retourner = lignes.aggregate(
        total=Sum('retours__quantite_retour')
    )['total'] or 0

    # --- calcul du total ---
    total = sum(Decimal(l.sous_total) for l in lignes)
    # -- Calcul du total quantite vendu
    total_quantite = sum(Decimal(l.quantite) for l in lignes)
    # -- Calcul du total de reduction
    total_reduction = sum(Decimal(l.montant_reduction) for l in lignes)

    # --- génération QR code ---
    try:
        qr_data = (
            f"Reçu Vente : {vente.code}\n"
            f"Date : {vente.date_vente.strftime('%d/%m/%Y %H:%M')}\n"
            f"Nombre d'articles : {lignes.count()}\n"
            f"Total : {total} GNF\n"
            f"Nom du Client : {vente.nom_complet_client}\n"
            f"Téléphone du Client : {vente.telclt_client}\n"
            f"Adresse du Client : {vente.adresseclt_client}\n"
        )

        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buffer = BytesIO()
        img.save(buffer, format="PNG")
        qr_code_base64 = base64.b64encode(buffer.getvalue()).decode()
    except Exception as e:
        qr_code_base64 = None
        messages.warning(request, f"QR code non généré : {e}")

    # --- contexte pour le template ---
    context = {
        "vente": vente,
        "lignes": lignes,
        "total": total,
        "today": now(),
        'has_retour' : has_retour,
        'total_quantite_retourner' : total_quantite_retourner,
        'total_quantite' : total_quantite,
        'total_reduction' : total_reduction,
        "qr_code_base64": qr_code_base64,
        "entreprise": Entreprise.objects.first(),  # Assure-toi qu'il y a bien une instance
    }

    return render(request, "gestion_produits/recu_ventes/recu_vente_global.html", context)

#================================================================================================
# Fonction pour afficher la listes des catégories
#================================================================================================
@login_required
def listes_categorie(request):
    try:
        # Récupérer toutes les catégories par ordre décroissant d'id
        listes_categories = CategorieProduit.objects.all().order_by('-id')

        # Pagination : 10 catégories par page
        paginator = Paginator(listes_categories, 10)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

        total_categories = listes_categories.count()
    except Exception as ex:
        messages.warning(request, f"Erreur lors du chargement des catégories : {str(ex)}")
        page_obj = []
        total_categories = 0

    context = {
        'liste_categories': page_obj,  # objet paginé pour le template
        'total_categories': total_categories,
    }
    return render(request, "gestion_produits/listes_categorie.html", context)

#================================================================================================
# Fonction pour modifier les informations d'une catégorie de produit
#================================================================================================
@login_required
def modifier_categorie(request):
    if request.method == 'POST':
        cat_id = request.POST.get('id_modif')
        nom = request.POST.get('nom_modif')
        description = request.POST.get('description_modif')

        if not cat_id or not nom:
            messages.error(request, "L'identifiant et le nom de la catégorie sont obligatoires.")
            return redirect('produits:listes_categorie')

        try:
            categorie = CategorieProduit.objects.get(id=cat_id)
            categorie.desgcategorie = nom
            categorie.description = description
            categorie.save(update_fields=['desgcategorie', 'description'])

            messages.success(request, "Catégorie modifiée avec succès !")
        except CategorieProduit.DoesNotExist:
            messages.error(request, "La catégorie spécifiée n'existe pas.")
        except Exception as ex:
            messages.error(request, f"Erreur lors de la modification : {str(ex)}")

        return redirect('produits:listes_categorie')

#================================================================================================
# Fonction pour supprimer une catégorie de produit
#================================================================================================
@login_required
def supprimer_categorie(request):
    if request.method == 'POST':
        cat_id = request.POST.get('id_supprime')
        if not cat_id:
            messages.error(request, "Aucun identifiant de catégorie fourni.")
            return redirect('produits:listes_categorie')

        try:
            categorie = CategorieProduit.objects.get(id=cat_id)

            # Vérifier si la catégorie est utilisée par un produit
            if Produits.objects.filter(categorie=cat_id).exists():
                messages.warning(
                    request,
                    "Impossible de supprimer cette catégorie car elle est utilisée par un produit. "
                    "Veuillez d'abord supprimer les produits associés."
                )
                return redirect('produits:listes_categorie')

            # Supprimer la catégorie
            categorie.delete()
            
            messages.success(request, "Catégorie supprimée avec succès !")

        except CategorieProduit.DoesNotExist:
            messages.error(request, "Catégorie introuvable.")
        except Exception as ex:
            messages.error(request, f"Erreur lors de la suppression : {str(ex)}")

        return redirect('produits:listes_categorie')

#================================================================================================
# Fonction pour supprimer un produit donné
#================================================================================================
@login_required
def supprimer_produits(request):
    if request.method != 'POST':
        messages.error(request, "Méthode invalide pour cette action.")
        return redirect('produits:listes_produits')

    prod_id = request.POST.get('id_supprimer')
    if not prod_id:
        messages.error(request, "Aucun produit sélectionné pour la suppression.")
        return redirect('produits:listes_produits')

    try:
        produit = Produits.objects.get(id=prod_id)

        # Vérifier les dépendances
        if LigneVente.objects.filter(produit=produit).exists():
            messages.warning(
                request,
                "Impossible de supprimer ce produit car il est utilisé dans une vente."
            )
            return redirect('produits:listes_produits')

        if StockProduit.objects.filter(produit=produit).exists():
            messages.warning(
                request,
                "Impossible de supprimer ce produit car il est utilisé dans un stock."
            )
            return redirect('produits:listes_produits')

        if Commandes.objects.filter(produits=produit).exists():
            messages.warning(
                request,
                "Impossible de supprimer ce produit car il est utilisé dans une commande."
            )
            return redirect('produits:listes_produits')

        if LivraisonsProduits.objects.filter(produits=produit).exists():
            messages.warning(
                request,
                "Impossible de supprimer ce produit car il est utilisé dans une livraison."
            )
            return redirect('produits:listes_produits')

        # ----- Suppression -----
        produit.delete()
        
        messages.success(request, "Produit supprimé avec succès !")
    except Produits.DoesNotExist:
        messages.error(request, "Produit introuvable.")
    except Exception as ex:
        messages.error(request, f"Erreur lors de la suppression : {str(ex)}")

    return redirect('produits:listes_produits')

#================================================================================================
# Fonction pour supprimer un produit donné
#================================================================================================
@login_required
def supprimer_produits_stock(request):
    if request.method != 'POST':
        messages.error(request, "Méthode invalide pour cette action.")
        return redirect('produits:listes_produits_stock')

    stock_id = request.POST.get('id_supprimer')
    if not stock_id:
        messages.error(request, "Aucun stock sélectionné pour la suppression.")
        return redirect('produits:listes_produits_stock')

    try:
        stock = StockProduit.objects.select_related('produit', 'entrepot', 'magasin').get(id=stock_id)

        # ----- Ancienne valeur pour audit -----
        ancienne_valeur = {
            "id_stock": stock.id,
            "produit": stock.produit.desgprod,
            "reference": stock.produit.refprod,
            "quantite": stock.qtestock,
            "seuil": stock.seuil,
            "entrepot": str(stock.entrepot) if stock.entrepot else "N/A",
            "magasin": str(stock.magasin) if stock.magasin else "N/A",
        }

        # ----- Suppression -----
        stock.delete()

        # ----- Audit -----
        enregistrer_audit(
            utilisateur=request.user,
            action="Suppression stock produit",
            table="StockProduit",
            ancienne_valeur=ancienne_valeur,
            nouvelle_valeur=None
        )

        # ----- Notification interne -----
        Notification.objects.create(
            destinataire=request.user,
            destinataire_email = settings.ADMIN_EMAIL,
            titre="🗑 Suppression de stock",
            message=(
                f"Le stock du produit {ancienne_valeur['produit']} "
                f"a été supprimé avec succès."
            )
        )

        # ----- Email admin -----
        try:
            sujet = "🗑 Suppression d’un stock produit"
            contenu = f"""
            Une suppression de stock a été effectuée.

            Utilisateur : {request.user.get_full_name()}
            Date : {timezone.now().strftime('%d/%m/%Y %H:%M')}

            Détails du stock supprimé :
            - Produit : {ancienne_valeur['produit']}
            - Référence : {ancienne_valeur['reference']}
            - Quantité : {ancienne_valeur['quantite']}
            - Seuil : {ancienne_valeur['seuil']}
            - Entrepôt : {ancienne_valeur['entrepot']}
            - Magasin : {ancienne_valeur['magasin']}
            """
            EmailMessage(
                sujet,
                contenu,
                settings.DEFAULT_FROM_EMAIL,
                [settings.ADMIN_EMAIL]
            ).send(fail_silently=False)
        except Exception as e:
            logger.error(f"Erreur envoi email suppression stock : {str(e)}")
            messages.warning(
                request,
                "Stock supprimé, mais l'email d'information n'a pas pu être envoyé."
            )

        messages.success(request, "Stock produit supprimé avec succès.")

    except StockProduit.DoesNotExist:
        messages.error(request, "Stock introuvable.")
    except Exception as ex:
        messages.error(request, f"Erreur lors de la suppression : {str(ex)}")

    return redirect('produits:listes_produits_stock')

#================================================================================================
# Fonction pour supprimer une commande donnée
#================================================================================================

@login_required
def supprimer_commandes(request):
    if request.method != 'POST':
        messages.error(request, "Méthode invalide pour cette action.")
        return redirect('produits:listes_des_commandes')

    commande_id = request.POST.get('id_supprimer')
    if not commande_id:
        messages.warning(request, "Aucune commande sélectionnée pour suppression.")
        return redirect('produits:listes_des_commandes')

    try:
        commande = get_object_or_404(Commandes, id=commande_id)

        # Vérifier si la commande est liée à des livraisons
        if commande.livraisonsproduits_set.exists():
            messages.warning(
                request,
                "Impossible de supprimer cette commande car elle est déjà liée à des livraisons."
            )
            return redirect('produits:listes_des_commandes')

        # ----- Suppression -----
        commande.delete()

        messages.success(request, "Commande supprimée avec succès ✔")

    except Exception as ex:
        messages.error(request, f"Erreur lors de la suppression : {str(ex)}")

    return redirect('produits:listes_des_commandes')

#================================================================================================
# Fonction pour supprimer une livraisons donnée
#================================================================================================

@login_required
def supprimer_livraisons(request):
    if request.method != 'POST':
        messages.error(request, "Méthode non autorisée.")
        return redirect('produits:listes_des_livraisons')

    livraison_id = request.POST.get('id_supprimer')

    if not livraison_id:
        messages.warning(request, "Aucune livraison sélectionnée.")
        return redirect('produits:listes_des_livraisons')

    try:
        with transaction.atomic():
            # 1️⃣ Récupérer la livraison
            livraison = get_object_or_404(LivraisonsProduits, id=livraison_id)
            produit = livraison.produits
            quantite = livraison.qtelivrer
            numlivrer = livraison.numlivrer

            # 2️⃣ Restaurer le stock produit
            stock_produit = StockProduit.objects.filter(produit=produit).first()
            if stock_produit:
                stock_produit.qtestock = stock_produit.qtestock - quantite  # <-- correction
                stock_produit.save(update_fields=['qtestock'])

            # 4️⃣ Supprimer la livraison
            livraison.delete()

        # 6️⃣ Notification interne
        Notification.objects.create(
            destinataire=request.user,
            destinataire_email = settings.ADMIN_EMAIL,
            titre="🗑 Suppression de livraison",
            message=(
                f"La livraison {numlivrer} du produit "
                f"{produit.desgprod} a été supprimée."
            )
        )

        # 7️⃣ Email administrateur
        try:
            sujet = "🗑 Suppression d'une livraison"
            contenu = f"""
            Une livraison a été supprimée.

            Numéro livraison : {numlivrer}
            Produit : {produit.desgprod}
            Quantité : {quantite}
            Utilisateur : {request.user}
            Date : {timezone.now().strftime('%d/%m/%Y %H:%M')}
            """
            email = EmailMessage(
                sujet,
                contenu,
                settings.DEFAULT_FROM_EMAIL,
                [settings.ADMIN_EMAIL]
            )
            email.send(fail_silently=False)
        except Exception as e:
            logger.error(f"Erreur email suppression livraison : {str(e)}")
            messages.warning(
                request,
                "Livraison supprimée mais l'email d'information n'a pas pu être envoyé."
            )

        messages.success(request, "Livraison supprimée avec succès. Stock mis à jour ✔")

    except Exception as ex:
        messages.error(
            request,
            f"Erreur lors de la suppression de la livraison : {str(ex)}"
        )

    return redirect('produits:listes_des_livraisons')

#================================================================================================
# Fonction pour supprimer une vente donnée
#================================================================================================
@login_required
def supprimer_ventes(request):
    if request.method != 'POST':
        messages.warning(request, "Méthode non autorisée.")
        return redirect('produits:listes_des_ventes')

    ligne_id = request.POST.get('id_supprimer')
    if not ligne_id:
        messages.warning(request, "Aucune vente sélectionnée.")
        return redirect('produits:listes_des_ventes')

    try:
        with transaction.atomic():

            # 1️⃣ Récupérer la ligne de vente
            ligne = get_object_or_404(LigneVente, id=ligne_id)
            vente = ligne.vente
            code_vente = vente.code

            # 2️⃣ Récupérer toutes les lignes de la vente
            lignes = LigneVente.objects.select_related('produit').filter(vente=vente)

            # 3️⃣ Restaurer le stock global
            for l in lignes:
                stock, created = StockProduit.objects.get_or_create(
                    produit=l.produit,
                    defaults={"qtestock": 0}
                )
                stock.qtestock += l.quantite
                stock.save(update_fields=["qtestock"])

            # 4️⃣ Audit
            ancienne_valeur = {
                "Vente": code_vente,
                "Produits": [
                    {
                        "Produit": l.produit.desgprod,
                        "Quantité": l.quantite,
                        "Sous-total": l.sous_total
                    } for l in lignes
                ],
                "Utilisateur": request.user.get_full_name(),
                "Date": timezone.now().strftime('%d/%m/%Y %H:%M')
            }

            enregistrer_audit(
                utilisateur=request.user,
                action="Suppression",
                table="VenteProduit",
                ancienne_valeur=ancienne_valeur,
                nouvelle_valeur=None
            )

            # 5️⃣ Supprimer lignes + vente
            lignes.delete()
            vente.delete()

            # 6️⃣ Notification
            Notification.objects.create(
                destinataire=request.user,
                destinataire_email = settings.ADMIN_EMAIL,
                titre=f"🗑 Suppression vente {code_vente}",
                message="La vente a été supprimée et le stock restauré."
            )

            # 7️⃣ Email admin
            try:
                EmailMessage(
                    subject=f"🗑 Suppression d'une vente - {code_vente}",
                    body=f"""
            Une vente a été supprimée.

            Code : {code_vente}
            Utilisateur : {request.user.get_full_name()}
            Date : {timezone.now().strftime('%d/%m/%Y %H:%M')}

            Le stock a été restauré automatiquement.
            """,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[settings.ADMIN_EMAIL]
                ).send()
            except Exception:
                messages.warning(request, "Email non envoyé.")

        messages.success(
            request,
            f"✅ Vente {code_vente} supprimée avec succès. Stock restauré ✔"
        )
    except Exception as e:
        messages.error(request, f"⚠️ Erreur suppression : {e}")

    return redirect('produits:listes_des_ventes')

#================================================================================================
# Fonction pour afficher la liste de tout les produits
#================================================================================================
@login_required
def listes_produits(request):
    try:
        # ================= LISTE DES PRODUITS =================
        produits = (
            Produits.objects
            .select_related('categorie')
            .order_by('desgprod')
        )
        total_quantite_restante = StockProduit.objects.aggregate(
            total=Sum('qtestock')
        )['total'] or 0
        total_produit = produits.count()

        # ================= TOTAL PAR CATÉGORIE =================
        total_par_categorie = (
            produits
            .values('categorie__desgcategorie')
            .annotate(
                nombre_produits=Count('id', distinct=True),
                quantite_stock=Sum('stocks__qtestock'),
                valeur_stock=Sum(F('stocks__qtestock') * F('pu'))
            )
            .order_by('categorie__desgcategorie')
        )

        listes_produits = pagination_liste(request, produits)

    except Exception as ex:
        messages.warning(request, f"Erreur de récupération des produits : {str(ex)}")
        listes_produits = []
        total_produit = 0
        total_quantite_restante = 0
        total_par_categorie = []

    context = {
        'listes_produits': listes_produits,
        'total_produit': total_produit,
        'total_par_categorie': total_par_categorie, 
        'total_quantite_restante' : total_quantite_restante,
    }

    return render(
        request,
        "gestion_produits/listes_produits.html",
        context
    )

#================================================================================================
# Fonction pour afficher la liste de tout les produits
#================================================================================================
@login_required
def listes_produits_stock(request):
    try:
        # ================= LISTE DES STOCKS =================
        listes_stock = (
            StockProduit.objects
            .select_related(
                'produit',
                'produit__categorie'
            )
            .order_by('produit__desgprod')
        )
        total_stocks = StockProduit.objects.aggregate(
            total=Sum('qtestock')
        )['total'] or 0
        total_produit = listes_stock.count()

        # ================= TOTAL PAR CATÉGORIE =================
        total_par_categorie = (
            listes_stock
            .values('produit__categorie__desgcategorie',
            )
            .annotate(
                nombre_produits=Count('id', distinct=True),
                quantite_restante=Sum('qtestock'),
                valeur_stock=Sum(F('qtestock') * F('produit__pu'))
            )
            .order_by('produit__categorie__desgcategorie')
        )
        
        listes_stock = pagination_liste(request, listes_stock)
    except Exception as ex:
        messages.error(
            request,
            f"Erreur de récupération des produits en stock : {str(ex)}"
        )
        return redirect('produits:listes_produits_stock')

    context = {
        'listes_produits': listes_stock,
        'total_produit': total_produit,
        'total_par_categorie': total_par_categorie,
        'total_stocks' : total_stocks,
    }

    return render(
        request,
        "gestion_produits/stocks/lites_produits_stocks.html",
        context
    )

#================================================================================================
# Fonction pour afficher la liste de tout les livraisons
#================================================================================================
@login_required
def listes_des_livraisons(request):

    try:
        # ================= LIVRAISONS =================
        livraisons_qs = LivraisonsProduits.objects.select_related(
            'commande', 'produits', 'produits__categorie'
        )

        # ================= QUANTITÉ LIVRÉE PAR COMMANDE =================
        livraison_par_commande = (
            LivraisonsProduits.objects
            .values('commande')
            .annotate(total_livree=Sum('qtelivrer'))
        )

        livraison_map = {
            l['commande']: l['total_livree'] for l in livraison_par_commande
        }

        # ================= LISTE + QTE RESTANTE =================
        listes_livraisons = livraisons_qs.order_by('-id')

        total_quantite_restante = 0
        for l in listes_livraisons:
            total_livree = livraison_map.get(l.commande_id, 0)
            l.qte_restante = max(l.commande.qtecmd - total_livree, 0)
            total_quantite_restante += l.qte_restante

        # ================= TOTAUX GLOBAUX =================
        total_livraison = listes_livraisons.count()

        total_quantite_livrer = listes_livraisons.aggregate(
            total=Sum('qtelivrer')
        )['total'] or 0
        
        total_qtecmd = listes_livraisons.aggregate(
            total=Sum('commande__qtecmd')
        )['total'] or 0

        total_quantite_livrer = listes_livraisons.aggregate(
            total=Sum('qtelivrer')
        )['total'] or 0

        total_quantite_restante = listes_livraisons.aggregate(
            total=Sum(F('commande__qtecmd') - F('qtelivrer'))
        )['total'] or 0

        # ================= TOTAUX PAR CATÉGORIE =================
        total_par_categorie = (
            LivraisonsProduits.objects
            .values('produits__categorie__desgcategorie')
            .annotate(
                nombre_livraisons=Count('id'),
                total_qtelivree=Sum('qtelivrer'),
                total_qtecmd=Sum('commande__qtecmd')
            )
            .annotate(
                total_qte_restante=F('total_qtecmd') - F('total_qtelivree')
            )
            .order_by('produits__categorie__desgcategorie')
        )
        # ================= TOTAUX PAR PRODUIT =================
        total_par_produit = (
            LivraisonsProduits.objects
            .values(
                'produits_id',
                'produits__desgprod',
                'produits__categorie__desgcategorie'
            )
            .annotate(
                nombre_livraisons=Count('id'),
                total_qtelivree=Sum('qtelivrer'),
                total_qtecmd=Sum('commande__qtecmd')
            )
            .annotate(
                total_qte_restante=F('total_qtecmd') - F('total_qtelivree')
            )
            .order_by('produits__desgprod')
        )

        # ================= PAGINATION =================
        listes_livraisons = pagination_liste(request, listes_livraisons)

    except Exception as ex:
        messages.warning(request, f"Erreur de récupération : {ex}")
        listes_livraisons = []
        total_livraison = 0
        total_quantite_livrer = 0
        total_quantite_restante = 0
        total_par_categorie = []
        total_par_produit = []
        total_qtecmd = 0
        total_quantite_livrer = 0
        total_quantite_restante = 0

    context = {
        'listes_livraisons': listes_livraisons,
        'total_livraison': total_livraison,
        'total_quantite_livrer': total_quantite_livrer,
        'total_quantite_restante': total_quantite_restante,
        'total_par_categorie': total_par_categorie,
        'total_par_produit': total_par_produit,
        'total_qtecmd' : total_qtecmd,
        
    }

    return render(
        request,
        "gestion_produits/livraisons/listes_livraisons.html",
        context
    )

#================================================================================================
# Fonction pour filtrer la liste des livraisons par date
#================================================================================================
@login_required
def filtrer_listes_livraisons(request):

    date_debut = request.GET.get("date_debut")
    date_fin = request.GET.get("date_fin")

    try:
        # ================= QUERYSET DE BASE =================
        livraisons_qs = LivraisonsProduits.objects.select_related(
            'commande',
            'produits',
            'produits__categorie'
        ).order_by('-id')

        # ================= FILTRE PAR DATE =================
        if date_debut and date_fin:
            livraisons_qs = livraisons_qs.filter(
                datelivrer__range=[date_debut, date_fin]
            )
        elif date_debut:
            livraisons_qs = livraisons_qs.filter(
                datelivrer=date_debut
            )
        elif date_fin:
            livraisons_qs = livraisons_qs.filter(
                datelivrer=date_fin
            )

        # ================= QTE LIVRÉE PAR COMMANDE =================
        livraison_par_commande = (
            livraisons_qs
            .values('commande_id')
            .annotate(total_livree=Sum('qtelivrer'))
        )

        livraison_map = {
            l['commande_id']: l['total_livree']
            for l in livraison_par_commande
        }

        # ================= QTE RESTANTE =================
        total_quantite_restante = 0
        for l in livraisons_qs:
            total_livree = livraison_map.get(l.commande_id, 0)
            l.qte_restante = max(l.commande.qtecmd - total_livree, 0)
            total_quantite_restante += l.qte_restante

        # ================= TOTAUX GLOBAUX =================
        total_livraison = livraisons_qs.count()

        total_qtecmd = livraisons_qs.aggregate(
            total=Sum('commande__qtecmd')
        )['total'] or 0

        total_quantite_livrer = livraisons_qs.aggregate(
            total=Sum('qtelivrer')
        )['total'] or 0

        total_quantite_restante = total_qtecmd - total_quantite_livrer

        # ================= TOTAUX PAR CATÉGORIE =================
        total_par_categorie = (
            livraisons_qs
            .values('produits__categorie__desgcategorie')
            .annotate(
                nombre_livraisons=Count('id'),
                total_qtecmd=Sum('commande__qtecmd'),
                total_qtelivree=Sum('qtelivrer'),
            )
            .annotate(
                total_qte_restante=F('total_qtecmd') - F('total_qtelivree')
            )
            .order_by('produits__categorie__desgcategorie')
        )

        # ================= TOTAUX PAR PRODUIT =================
        total_par_produit = (
            livraisons_qs
            .values(
                'produits_id',
                'produits__desgprod',
                'produits__categorie__desgcategorie'
            )
            .annotate(
                nombre_livraisons=Count('id'),
                total_qtecmd=Sum('commande__qtecmd'),
                total_qtelivree=Sum('qtelivrer'),
            )
            .annotate(
                total_qte_restante=F('total_qtecmd') - F('total_qtelivree')
            )
            .order_by('produits__desgprod')
        )

        # ================= PAGINATION (FIN) =================
        listes_livraisons_filtre = pagination_liste_filtre(
            request,
            livraisons_qs
        )
                # ================= QUERY PARAMS (IMPORTANT) =================
        query_params = request.GET.copy()
        query_params.pop('page', None)
        query_params = urlencode(query_params)

    except Exception as ex:
        messages.warning(request, f"Erreur lors du filtrage : {ex}")
        listes_livraisons_filtre = []
        total_livraison = total_qtecmd = total_quantite_livrer = total_quantite_restante = 0
        total_par_categorie = []
        total_par_produit = []
        query_params = ""

    return render(
        request,
        "gestion_produits/livraisons/listes_livraisons.html",
        {
            "date_debut": date_debut,
            "date_fin": date_fin,
            "listes_livraisons_filtre": listes_livraisons_filtre,
            "total_livraison": total_livraison,
            "total_qtecmd": total_qtecmd,
            "total_quantite_livrer": total_quantite_livrer,
            "total_quantite_restante": total_quantite_restante,
            "total_par_categorie": total_par_categorie,
            "total_par_produit": total_par_produit,
            
            'query_params' : query_params,
        }
    )

#================================================================================================
# Fonction pour afficher la liste des ventes
#================================================================================================
@login_required
def listes_des_ventes(request):
    try:
        # ================= LIGNES DE VENTE =================
        lignes = (
            LigneVente.objects
            .select_related('vente', 'produit', 'produit__categorie')
            .order_by('-id')
        )
        total_vendus = LigneVente.objects.aggregate(
            total=Sum('quantite')
        )['total'] or 0
        
        
        total_ventes = lignes.count()
        total_montant_ventes = 0
        benefice_global = 0
        total_retourner = 0
        listes_ventes = []

        for ligne in lignes:

            # Mise à jour des totaux
            benefice_global += ligne.benefice
            total_montant_ventes += ligne.sous_total
            total_retourner += ligne.quantite_retournee

            listes_ventes.append(ligne)
            
        # ================= TOTAL PAR CATÉGORIE =================
        total_par_categorie = (
            lignes
            .values('produit__categorie__desgcategorie')
            .annotate(
                total_montant=Sum('sous_total'),
                total_quantite=Sum('quantite'),
                total_vendu=Sum('quantite'),
            )
            .order_by('produit__categorie__desgcategorie')
        )
        #=================  Total par produit ================= 
        total_par_produit = (
            lignes
            .values('produit__desgprod')
            .annotate(
                    total_quantite=Sum('quantite'),
                    total_montant=Sum('sous_total')
                )
                .order_by('produit__desgprod')
            )

        listes_ventes = pagination_liste(request, listes_ventes)

    except Exception as ex:
        messages.warning(request, f"Erreur de récupération des ventes : {str(ex)}")
        listes_ventes = []
        total_ventes = 0
        total_montant_ventes = 0
        benefice_global = 0
        total_vendus = 0
        total_retourner = 0
        total_par_categorie = []
        total_par_produit = []

    # ================= CONTEXT =================
    context = {
        'listes_ventes': listes_ventes,
        'total_ventes': total_ventes,
        'total_montant_ventes': total_montant_ventes,
        'benefice_global': benefice_global,
        'total_par_categorie': total_par_categorie,
        'total_vendus' : total_vendus,
        'total_retourner' : total_retourner,
        'total_par_produit' : total_par_produit,
    }
    return render(
        request,"gestion_produits/ventes/listes_ventes.html",context)

#================================================================================================
# Fonction pour afficher la liste des commandes éffectuées
#================================================================================================
@login_required
def listes_des_commandes(request):
    total_commande = 0
    listes_commandes = []
    total_par_categorie = []
    total_par_produit = []
    total_quantite = 0
    try:
        # ------------------ LISTE DES COMMANDES ------------------
        listes_commandes_qs = Commandes.objects.select_related(
            'produits', 'produits__categorie'
        ).order_by('-id')
        total_commande = listes_commandes_qs.count()

        # ------------------ TOTAL PAR CATEGORIE ------------------
        total_par_categorie = (
            listes_commandes_qs
            .values('produits__categorie__desgcategorie')
            .annotate(
                nombre_commandes=Count('id', distinct=True),
                total_quantite=Sum('qtecmd'),
                valeur_commandes=Sum(F('qtecmd') * F('produits__pu'))
            )
            .order_by('produits__categorie__desgcategorie')
        )

        # ------------------ TOTAL PAR PRODUIT ------------------
        total_par_produit = (
            listes_commandes_qs
            .values('produits__categorie__desgcategorie', 'produits__refprod', 'produits__desgprod')
            .annotate(
                nombre_commandes=Count('id', distinct=True),
                total_quantite=Sum('qtecmd'),
                valeur_commandes=Sum(F('qtecmd') * F('produits__pu'))
            )
            .order_by('produits__categorie__desgcategorie', 'produits__refprod')
        )

        # ------------------ TOTAL GLOBAUX ------------------
        total_quantite = listes_commandes_qs.aggregate(total_qte=Sum('qtecmd'))['total_qte'] or 0

        # ------------------ PAGINATION ------------------
        if 'pagination_lis' in globals():
            listes_commandes = pagination_liste(request, listes_commandes_qs)
        else:
            listes_commandes = listes_commandes_qs

    except Exception as ex:
        messages.warning(request, f"Erreur de récupération des commandes : {str(ex)} !")
        listes_commandes = []
        total_commande = 0
        total_par_produit = []
        total_par_categorie = []
        total_quantite = 0

    context = {
        'listes_commandes': listes_commandes,
        'total_commande': total_commande,
        'total_par_categorie': total_par_categorie,
        'total_par_produit': total_par_produit,
        'total_quantite': total_quantite,
    }

    return render(request, "gestion_produits/commandes/listes_commandes.html", context)

#================================================================================================
# Fonction pour filter la liste des commandes selon un intervalle de date donnée
#================================================================================================
@login_required
def filtrer_listes_commandes(request):
    """
    Filtre les commandes selon la date
    (DateField : une date ou intervalle)
    """
    date_debut = request.GET.get("date_debut")
    date_fin = request.GET.get("date_fin")

    total_commande = 0
    total_par_categorie = []
    total_par_produit = []
    listes_commandes_filtre = []

    try:
        # ================== QUERYSET DE BASE ==================
        commande_qs = Commandes.objects.select_related(
            'produits', 'produits__categorie'
        ).order_by("-datecmd")

        # ================== FILTRE PAR DATE ==================
        if date_debut and date_fin:
            commande_qs = commande_qs.filter(datecmd__range=[date_debut, date_fin])
        elif date_debut:
            commande_qs = commande_qs.filter(datecmd=date_debut)
        elif date_fin:
            commande_qs = commande_qs.filter(datecmd=date_fin)

        # ================== TOTAL DES COMMANDES ==================
        total_commande = commande_qs.count()

        # ================== TOTAL PAR CATEGORIE ==================
        total_par_categorie = commande_qs.values(
            'produits__categorie__desgcategorie'
        ).annotate(
            nombre_commandes=Count('id', distinct=True),
            total_quantite=Sum('qtecmd'),
            valeur_commandes=Sum(F('qtecmd') * F('produits__pu'))
        ).order_by('produits__categorie__desgcategorie')

        # ================== TOTAL PAR PRODUIT ==================
        total_par_produit = commande_qs.values(
            'produits__categorie__desgcategorie',
            'produits__refprod',
            'produits__desgprod'
        ).annotate(
            nombre_commandes=Count('id', distinct=True),
            total_quantite=Sum('qtecmd'),
            valeur_commandes=Sum(F('qtecmd') * F('produits__pu'))
        ).order_by('produits__categorie__desgcategorie', 'produits__refprod')

        # ================== TOTAL GLOBAUX ==================
        total_quantite = commande_qs.aggregate(total_qte=Sum('qtecmd'))['total_qte'] or 0

        # ================== PAGINATION ==================
        if 'pagination_liste' in globals():
            listes_commandes_filtre = pagination_liste_filtre(request, commande_qs)
        else:
            listes_commandes_filtre = commande_qs
        
        # ================= QUERY PARAMS (IMPORTANT) =================
        query_params = request.GET.copy()
        query_params.pop('page', None)
        query_params = urlencode(query_params)

    except Exception as ex:
        messages.warning(request, f"Erreur lors du filtrage des commandes : {str(ex)}")
        listes_commandes_filtre = []
        total_commande = 0
        total_par_categorie = []
        total_par_produit = []
        total_quantite = 0
        query_params = ""

    context = {
        "date_debut": date_debut,
        "date_fin": date_fin,
        "listes_commandes_filtre": listes_commandes_filtre,
        "total_commande": total_commande,
        "total_par_categorie": total_par_categorie,
        "total_par_produit": total_par_produit,
        "total_quantite": total_quantite,
        'query_params' : query_params,
    }

    return render(request, "gestion_produits/commandes/listes_commandes.html", context)

#================================================================================================
# Fonction pour filter la liste des vente selon un intervalle de date donnée
#================================================================================================
@login_required
def filtrer_listes_ventes(request):
    """
    Filtre les ventes selon la date
    + statistiques
    + pagination AVEC conservation des paramètres GET
    """

    date_debut = request.GET.get("date_debut")
    date_fin = request.GET.get("date_fin")

    try:
        # ================== QUERYSET DE BASE ==================
        ventes_qs = (
            LigneVente.objects
            .select_related('produit', 'vente', 'produit__categorie')
            .order_by("-vente__date_vente")
        )

        # ================== FILTRE PAR DATE ==================
        if date_debut and date_fin:
            ventes_qs = ventes_qs.filter(
                date_saisie__range=[date_debut, date_fin]
            )

        # ================== STATISTIQUES GLOBALES ==================
        stats = ventes_qs.aggregate(
            total_vendus=Sum('quantite'),
            total_montant_ventes=Sum('sous_total'),
            benefice_global=Sum('benefice'),
            total_retourner=Sum('retours')
        )

        total_ventes = ventes_qs.count()
        total_vendus = stats['total_vendus'] or 0
        total_montant_ventes = stats['total_montant_ventes'] or 0
        benefice_global = stats['benefice_global'] or 0
        total_retourner = stats['total_retourner'] or 0

        # ================== PAR CATÉGORIE ==================
        total_par_categorie = (
            ventes_qs
            .values('produit__categorie__desgcategorie')
            .annotate(
                total_quantite=Sum('quantite'),
                total_montant=Sum('sous_total')
            )
            .order_by('produit__categorie__desgcategorie')
        )

        # ================== PAR PRODUIT ==================
        total_par_produit = (
            ventes_qs
            .values('produit__desgprod')
            .annotate(
                total_quantite=Sum('quantite'),
                total_montant=Sum('sous_total')
            )
            .order_by('produit__desgprod')
        )

        # ================== PAGINATION ==================
        listes_ventes_filtre = pagination_liste_filtre(request, ventes_qs)

        # ================== QUERY PARAMS POUR PAGINATION ==================
        query_params = request.GET.copy()
        query_params.pop('page', None)

    except Exception as ex:
        messages.warning(
            request,
            f"Erreur lors du filtrage des ventes : {str(ex)}"
        )

        listes_ventes_filtre = []
        total_ventes = 0
        total_vendus = 0
        total_montant_ventes = 0
        benefice_global = 0
        total_retourner = 0
        total_par_categorie = []
        total_par_produit = []
        query_params = ""

    # ================== CONTEXT ==================
    context = {
        "date_debut": date_debut,
        "date_fin": date_fin,

        "listes_ventes_filtre": listes_ventes_filtre,
        "total_ventes": total_ventes,
        "total_vendus": total_vendus,
        "total_montant_ventes": total_montant_ventes,
        "benefice_global": benefice_global,
        "total_retourner": total_retourner,

        "total_par_categorie": total_par_categorie,
        "total_par_produit": total_par_produit,

        # 🔑 CLÉ DE LA CORRECTION
        "query_params": query_params.urlencode() if query_params else "",
    }

    return render(
        request,
        "gestion_produits/ventes/listes_ventes.html",
        context
    )

#================================================================================================
# Fonction pour consulter un produit donnée
#================================================================================================
@login_required
def consulter_produit(request, id):
    try:
        produit = Produits.objects.get(id=id)
    except Produits.DoesNotExist:
        messages.error(request, "Produit introuvable.")
        return redirect('produits:listes_produits')

    context = {
        'produit': produit,
    }
    return render(request, 'gestion_produits/consulter_informations_eleves.html', context)

#================================================================================================
# Fonction pour editer un produit pour provoir le modifier
#================================================================================================
@login_required
def editer_produit(request, id):
    try:
        produit = Produits.objects.get(id=id)
    except Produits.DoesNotExist:
        messages.error(request, "Produit introuvable.")
        return redirect('produits:listes_produits')

    context = {
        'produit': produit,
        'categories': CategorieProduit.objects.all(),
    }
    return render(request, 'gestion_produits/modifier_produits.html', context)

#================================================================================================
# Fonction pour modifier un produit donnée
#================================================================================================
@login_required
def modifier_produit(request, id):
    try:
        produit = Produits.objects.get(id=id)
    except Produits.DoesNotExist:
        messages.error(request, "Produit introuvable.")
        return redirect('produits:editer_produit')

    categories = CategorieProduit.objects.all()

    if request.method == 'POST':
        produit.desgprod = request.POST.get('desgprod')
        produit.qtestock = request.POST.get('qtestock')
        produit.seuil = request.POST.get('seuil')
        produit.pu = request.POST.get('pu')
        produit.prix_en_gros = request.POST.get('prix_en_gros')

        categorie_id = request.POST.get('categorie')
        produit.categorie_id = categorie_id

        if 'photoprod' in request.FILES:
            produit.photoprod = request.FILES['photoprod']

        produit.save()

        messages.success(request, "Produit modifié avec succès !")
        return redirect('produits:listes_produits')

    context = {
        'produit': produit,
        'categories': categories,
    }
    return render(request, 'gestion_produits/modifier_produits.html', context)

#================================================================================================
# Fonction gérer les réferenes de produit
#================================================================================================
def generate_references(prefix, date_str, numero):
    return f"{prefix}{date_str}{str(numero).zfill(4)}"

#================================================================================================
# Fonction pour ajouter un nouveau produit
#================================================================================================
@login_required(login_url='gestionUtilisateur:connexion_utilisateur')
def nouveau_produit(request):
    prefix = "PROD"
    date_str = datetime.now().strftime("%Y%m%d")

    # Trouver le dernier produit du jour pour incrémentation
    last_produit = Produits.objects.filter(
        refprod__startswith=f"{prefix}{date_str}"
    ).order_by('-refprod').first()

    if last_produit:
        dernier_numero = int(last_produit.refprod[-4:])
    else:
        dernier_numero = 0

    # Première référence à afficher
    ref_generee = generate_references(prefix, date_str, dernier_numero + 1)

    # ------- TRAITEMENT DU FORMULAIRE -------
    if request.method == 'POST':

        refs = request.POST.getlist("refprod[]")
        noms = request.POST.getlist("desgprod[]")
        pus = request.POST.getlist("pu[]")
        pu_engros = request.POST.getlist("pu_engros[]")
        categories = request.POST.getlist("categorie[]")
        photos = request.FILES.getlist("photoprod[]")

        total = len(noms)
        success_count = 0

        # Vérification cohérence des listes
        if not (len(refs) == len(noms) == len(pus) == len(pu_engros) == len(categories)):
            messages.error(request, "Erreur : Données incomplètes dans le formulaire.")
            return redirect("produits:nouveau_produit")

        for i in range(total):
            ref = refs[i]
            desg = noms[i]
            try:
                pu = int(pus[i])
                pu_engro_val = int(pu_engros[i])
            except ValueError:
                messages.error(request, f"Erreur : Le prix doit être un nombre pour le produit {desg}.")
                return redirect("produits:nouveau_produit")

            cat_id = categories[i]
            photo = photos[i] if i < len(photos) else None

            # Vérifier doublons
            if Produits.objects.filter(refprod=ref).exists():
                messages.error(request, f"La Référence {ref} existe déjà.")
                return redirect('produits:nouveau_produit')
            elif Produits.objects.filter(desgprod=desg).exists():
                messages.error(request, f"Le nom du Produit {desg} existe déjà.")
                return redirect('produits:nouveau_produit')

            # Création du produit
            try:
                Produits.objects.create(
                    refprod=ref,
                    desgprod=desg,
                    pu=pu,
                    prix_en_gros=pu_engro_val,
                    photoprod=photo,
                    categorie_id=cat_id
                )
                success_count += 1
            except Exception as e:
                messages.error(request, f"Erreur lors de l’enregistrement de {ref} : {e}")

        if success_count > 0:
            messages.success(request, f"{success_count} produit(s) enregistré(s) avec succès.")
            return redirect("produits:listes_produits")

    # ------- CONTEXTE POUR LE TEMPLATE -------
    context = {
        "ref_generee": ref_generee,
        "categorie_choices": CategorieProduit.objects.all(),
    }

    return render(request, "gestion_produits/nouveau_produit.html", context)

#================================================================================================
# Fonction pour ajouter un nouveau produit
#================================================================================================
@login_required(login_url='gestionUtilisateur:connexion_utilisateur')
def ajouter_stock_multiple(request):
    produits = Produits.objects.all().order_by('desgprod')
    total_quantite = StockProduit.objects.aggregate(
        total = Sum('qtestock'))['total'] or 0

    if request.method == "POST":
        produit_ids = request.POST.getlist("produit[]")
        qte_list = request.POST.getlist("qtestock[]")
        seuil_list = request.POST.getlist("seuil[]")

        success_count = 0

        for i in range(len(produit_ids)):
            try:
                produit = Produits.objects.get(id=int(produit_ids[i]))

                qte = int(qte_list[i])
                seuil = int(seuil_list[i])

                # Création ou mise à jour du stock unique
                if qte != 0 :
                    stock, created = StockProduit.objects.get_or_create(
                        produit=produit,
                        defaults={
                            "qtestock": qte,
                            "seuil": seuil
                        }
                    )
                    if not stock.seuil :
                        if not created :
                            stock.qtestock += qte
                            stock.seuil = seuil
                            stock.save()
                    success_count += 1

            except Produits.DoesNotExist:
                messages.error(
                    request,
                    f"Produit introuvable à la ligne {i + 1}."
                )

            except ValueError as ve:
                messages.error(
                    request,
                    f"Quantité ou seuil invalide pour le produit sélectionné {str(ve)}."
                )
            except Exception as e:
                messages.error(
                    request,
                    f"Erreur pour le produit {produit.refprod} : {str(e)}"
                )

        messages.success(
            request,
            f"{success_count} produit(s) enregistré(s) / mis à jour avec succès."
        )

        return redirect("produits:listes_produits_stock")

    return render(
        request,
        "gestion_produits/stocks/ajouter_stock_multiple.html",
        {
            "produits": produits,
            'total_quantite' : total_quantite,
        }
    )


#================================================================================================
# Fonction pour imprimer la listes des produits
#================================================================================================
@login_required
def listes_produits_impression(request):

    listes_produits = Produits.objects.all().order_by('desgprod')

    total_quantite_restante = StockProduit.objects.aggregate(
            total=Sum('qtestock')
        )['total'] or 0
    total_produit = listes_produits.count()
    
        # ================= TOTAL PAR CATÉGORIE =================
    total_par_categorie = (
        listes_produits
            .values('categorie__desgcategorie')
            .annotate(
                nombre_produits=Count('id', distinct=True),
                quantite_stock=Sum('stocks__qtestock'),
            )
            .order_by('categorie__desgcategorie')
        )

    nom_entreprise = Entreprise.objects.first()
    context = {
        'nom_entreprise': nom_entreprise,
        'today': timezone.now(),
        'listes_produits' : listes_produits,
        'total_par_categorie' : total_par_categorie,
        'total_produit' : total_produit,
        'total_quantite_restante' : total_quantite_restante,
    }
    return render(
        request,
        'gestion_produits/impression_listes/apercue_avant_impression_listes_produits.html',
        context
    )

#================================================================================================
# Fonction pour imprimer la listes des Catégories Produits
#================================================================================================
@login_required
def listes_categorie_produits_impression(request):
    listes_categorie_produits = []
    try:
        listes_categorie_produits = CategorieProduit.objects.all()
    except Exception as ex:
        messages.warning(request, f"Erreur de récupération des données {str(ex)}")
    nom_entreprise = Entreprise.objects.first()
    context = {
        'nom_entreprise': nom_entreprise,
        'today': timezone.now(),
        'listes_categorie_produits' : listes_categorie_produits,
    }
    return render(
        request,
        'gestion_produits/impression_listes/apercue_avant_impression_listes_categorieproduits.html',
        context
    )

#================================================================================================
# Fonction pour afficher formulaire de choix de dates de saisie pour l'impression
#================================================================================================
@login_required
def choix_par_dates_ventes_impression(request):
    return render(request, 'gestion_produits/impression_listes/fiches_choix_impression_ventes.html')

#================================================================================================
# Fonction pour imprimer la listes des ventes
#================================================================================================
@login_required
def listes_ventes_impression(request):
    # Récupération des dates depuis POST
    date_debut = request.POST.get('date_debut')
    date_fin = request.POST.get('date_fin')

    lignes = LigneVente.objects.none()
    total_par_categorie = []
    total_par_produit = []
    total_quantite_produits = 0
    total_montant_produits = 0
    total_quantite_categories = 0
    total_montant_categories = 0
    total_quantite_retourner = 0
    benefice_global = 0

    if date_debut and date_fin:
        try:
            lignes = (
                LigneVente.objects
                .select_related('vente', 'produit', 'produit__categorie')
                .filter(date_saisie__range=[date_debut, date_fin])
                .order_by('-id')
            )
            # ------- Total par catégorie -------
            total_par_categorie = (
                lignes
                .values('produit__categorie__desgcategorie')
                .annotate(
                    total_montant=Sum('sous_total'),
                    total_quantite=Sum('quantite')
                )
                .order_by('produit__categorie__desgcategorie')
            )

            # Totaux globaux par catégorie
            total_quantite_categories = sum(c['total_quantite'] for c in total_par_categorie)
            total_montant_categories = sum(c['total_montant'] for c in total_par_categorie)

            # Total par produit
            total_par_produit = (
                lignes
                .values('produit__desgprod')
                .annotate(
                    total_montant=Sum('sous_total'),
                    total_quantite=Sum('quantite')
                )
                .order_by('produit__desgprod')
            )

            # Totaux globaux par produit
            total_quantite_produits = sum(p['total_quantite'] for p in total_par_produit)
            
            total_montant_produits = sum(p['total_montant'] for p in total_par_produit)

        except Exception as ex:
            messages.warning(request, f"Erreur lors de la récupération des ventes : {str(ex)}")

        # Regrouper les lignes par vente
        ventes_dict = {}
        for ligne in lignes:
            code_vente = ligne.vente.code
            if code_vente not in ventes_dict:
                ventes_dict[code_vente] = {
                    'vente': ligne.vente,
                    'lignes': [],
                    'total_vente': 0,
                    'total_quantite_vente': 0,  
                    'total_quantite_retourner': 0,  
                    'benefice_vente': 0,
                    'montant_reduction' : 0
                }

            ventes_dict[code_vente]['lignes'].append(ligne)
            ventes_dict[code_vente]['benefice_vente'] += ligne.benefice
            ventes_dict[code_vente]['total_vente'] += ligne.sous_total
            ventes_dict[code_vente]['montant_reduction'] += ligne.montant_reduction
            ventes_dict[code_vente]['total_quantite_vente'] += ligne.quantite  
            ventes_dict[code_vente]['total_quantite_retourner'] += ligne.quantite_retournee or 0

            benefice_global += ligne.benefice
            total_quantite_retourner += ligne.quantite_retournee

    ventes_liste = list(ventes_dict.values())
    nom_entreprise = Entreprise.objects.first()  # Si plusieurs, prendre le premier

    context = {
        'nom_entreprise': nom_entreprise,
        'today': timezone.now(),
        'ventes_liste': ventes_liste,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'benefice_global': benefice_global,
        'total_par_categorie': total_par_categorie,
        'total_par_produit': total_par_produit,
        'total_quantite_produits': total_quantite_produits,
        'total_montant_produits': total_montant_produits,
        'total_quantite_categories': total_quantite_categories,
        'total_montant_categories': total_montant_categories,
        'total_quantite_retourner' : total_quantite_retourner,
    }

    return render(
        request,
        'gestion_produits/impression_listes/apercue_avant_impression_listes_ventes.html',
        context
    )

#================================================================================================
# Fonction pour afficher le formulaire de choix de dates de saisie pour l'impression des Commandes
#================================================================================================
@login_required
def choix_par_dates_commandes_impression(request):
    return render(request, 'gestion_produits/impression_listes/fiches_choix_impression_commandes.html')

#================================================================================================
# Fonction pour imprimer la listes des Commandes
#================================================================================================
@login_required
def listes_commandes_impression(request):

    date_debut = request.POST.get('date_debut')
    date_fin = request.POST.get('date_fin')

    try:
        # ================= COMMANDES FILTRÉES =================
        listes_commandes = Commandes.objects.select_related(
            'produits',
            'produits__categorie'
        ).filter(
            datecmd__range=[date_debut, date_fin]
        ).order_by('-datecmd')

        # ================= TOTAUX GLOBAUX =================
        total_commandes = listes_commandes.count()

        total_quantite = listes_commandes.aggregate(
            total=Sum('qtecmd')
        )['total'] or 0

        total_valeur = listes_commandes.aggregate(
            total=Sum(
                ExpressionWrapper(
                    F('qtecmd') * F('produits__pu'),
                    output_field=IntegerField()
                )
            )
        )['total'] or 0

        # ================= TOTAL PAR CATÉGORIE =================
        total_par_categorie = (
            listes_commandes
            .values('produits__categorie__desgcategorie')
            .annotate(
                nombre_commandes=Count('id'),
                total_quantite=Sum('qtecmd'),
                valeur_commandes=Sum(
                    ExpressionWrapper(
                        F('qtecmd') * F('produits__pu'),
                        output_field=IntegerField()
                    )
                )
            )
            .order_by('produits__categorie__desgcategorie')
        )

        # ================= TOTAL PAR PRODUIT =================
        total_par_produit = (
            listes_commandes
            .values(
                'produits__refprod',
                'produits__desgprod',
                'produits__categorie__desgcategorie'
            )
            .annotate(
                nombre_commandes=Count('id'),
                total_quantite=Sum('qtecmd'),
                valeur_commandes=Sum(
                    ExpressionWrapper(
                        F('qtecmd') * F('produits__pu'),
                        output_field=IntegerField()
                    )
                )
            )
            .order_by('produits__desgprod')
        )

    except Exception as ex:
        messages.warning(request, f"Erreur impression commandes : {ex}")
        listes_commandes = []
        total_par_categorie = []
        total_par_produit = []
        total_commandes = 0
        total_quantite = 0
        total_valeur = 0

    nom_entreprise = Entreprise.objects.first()

    context = {
        'nom_entreprise': nom_entreprise,
        'today': timezone.now(),
        'date_debut': date_debut,
        'date_fin': date_fin,

        # données
        'listes_commandes': listes_commandes,
        'total_par_categorie': total_par_categorie,
        'total_par_produit': total_par_produit,

        # totaux globaux
        'total_commandes': total_commandes,
        'total_quantite': total_quantite,
        'total_valeur': total_valeur,
    }
    return render(
        request,
        'gestion_produits/impression_listes/apercue_avant_impression_listes_commandes.html',
        context)
    
#================================================================================================
# Fonction pour imprimer la listes des Produits en Stocks
#================================================================================================
@login_required
def listes_stocks_impression(request):

    listes_produits = StockProduit.objects.all().order_by('produit__desgprod')
    total_stocks = StockProduit.objects.aggregate(
        total=Sum('qtestock')
        )['total'] or 0
    total_produit = listes_produits.count()
    
        # ================= TOTAL PAR CATÉGORIE =================
    total_par_categorie = (
        listes_produits
        .values('produit__categorie__desgcategorie',)
        .annotate(
            nombre_produits=Count('id', distinct=True),
            quantite_restante=Sum('qtestock'),
            valeur_stock=Sum(F('qtestock') * F('produit__pu'))
        )
        .order_by('produit__categorie__desgcategorie')
        )

    # ================= CONTEXT =================
    nom_entreprise = Entreprise.objects.first()

    context = {
        'nom_entreprise': nom_entreprise,
        'today': timezone.now(),
        'listes_produits': listes_produits,
        'total_par_categorie' : total_par_categorie,
        'total_produit' : total_produit,
        'total_stocks' : total_stocks,
    }

    return render(
        request,
        'gestion_produits/impression_listes/stock/apercue_avant_impression_listes_stocks.html',
        context
    )

#================================================================================================
# Fonction pour afficher le formulaire de choix de dates de saisie pour l'impression des Livraisons
#================================================================================================
@login_required
def choix_par_dates_livraisons_impression(request):
    return render(request, 'gestion_produits/impression_listes/fiches_choix_impression_livraisons.html')

#================================================================================================
# Fonction pour imprimer la listes des Livraisons
#================================================================================================
@login_required
def listes_livraisons_impression(request):

    try:
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')

        # ================= QUERYSET DE BASE =================
        livraisons_qs = LivraisonsProduits.objects.select_related(
            'commande',
            'produits',
            'produits__categorie'
        )

        # ================= FILTRE PAR DATE =================
        if date_debut and date_fin:
            livraisons_qs = livraisons_qs.filter(
                datelivrer__range=[date_debut, date_fin]
            )
        elif date_debut:
            livraisons_qs = livraisons_qs.filter(datelivrer=date_debut)
        elif date_fin:
            livraisons_qs = livraisons_qs.filter(datelivrer=date_fin)

        livraisons_qs = livraisons_qs.order_by('-id')

        # ================= QTE LIVRÉE PAR COMMANDE =================
        livraison_par_commande = (
            livraisons_qs
            .values('commande')
            .annotate(total_livree=Sum('qtelivrer'))
        )

        livraison_map = {
            l['commande']: l['total_livree'] for l in livraison_par_commande
        }

        # ================= LISTE + QTE RESTANTE =================
        total_quantite_restante = 0
        for l in livraisons_qs:
            total_livree = livraison_map.get(l.commande_id, 0)
            l.qte_restante = max(l.commande.qtecmd - total_livree, 0)
            total_quantite_restante += l.qte_restante

        # ================= TOTAUX GLOBAUX =================
        total_livraison = livraisons_qs.count()

        total_qtecmd = livraisons_qs.aggregate(
            total=Sum('commande__qtecmd')
        )['total'] or 0

        total_quantite_livrer = livraisons_qs.aggregate(
            total=Sum('qtelivrer')
        )['total'] or 0

        total_quantite_restante = livraisons_qs.aggregate(
            total=Sum(F('commande__qtecmd') - F('qtelivrer'))
        )['total'] or 0

        # ================= TOTAUX PAR CATÉGORIE =================
        total_par_categorie = (
            livraisons_qs
            .values('produits__categorie__desgcategorie')
            .annotate(
                nombre_livraisons=Count('id'),
                total_qtecmd=Sum('commande__qtecmd'),
                total_qtelivree=Sum('qtelivrer')
            )
            .annotate(
                total_qte_restante=F('total_qtecmd') - F('total_qtelivree')
            )
            .order_by('produits__categorie__desgcategorie')
        )

        # ================= TOTAUX PAR PRODUIT =================
        total_par_produit = (
            livraisons_qs
            .values(
                'produits_id',
                'produits__desgprod',
                'produits__refprod',
                'produits__categorie__desgcategorie'
            )
            .annotate(
                nombre_livraisons=Count('id'),
                total_qtecmd=Sum('commande__qtecmd'),
                total_qtelivree=Sum('qtelivrer')
            )
            .annotate(
                total_qte_restante=F('total_qtecmd') - F('total_qtelivree')
            )
            .order_by('produits__desgprod')
        )

    except Exception as ex:
        messages.warning(request, f"Erreur impression livraisons : {ex}")
        livraisons_qs = []
        total_livraison = 0
        total_qtecmd = 0
        total_quantite_livrer = 0
        total_quantite_restante = 0
        total_par_categorie = []
        total_par_produit = []

    # ================= CONTEXTE =================
    nom_entreprise = Entreprise.objects.first()

    context = {
        'nom_entreprise': nom_entreprise,
        'today': timezone.now(),

        'listes_livraisons': livraisons_qs,
        'date_debut': date_debut,
        'date_fin': date_fin,

        # Totaux
        'total_livraison': total_livraison,
        'total_qtecmd': total_qtecmd,
        'total_quantite_livrer': total_quantite_livrer,
        'total_quantite_restante': total_quantite_restante,

        # Récaps
        'total_par_categorie': total_par_categorie,
        'total_par_produit': total_par_produit,
    }

    return render(
        request,
        'gestion_produits/impression_listes/apercue_avant_impression_listes_livraisons.html',
        context
    )

#================================================================================================
# Fonction pour afficher le formulaire de formulaire d'exportation des données
#================================================================================================
@login_required
def confirmation_exportation_vente(request):
    
    return render(request, 'gestion_produits/exportation/confirmation_exportation_ventes.html')

#=============================================================================================
# Fonction pour exporter les données des ventes
#==============================================================================================
@login_required
def export_ventes_excel_complet(request):
    """
    Export Excel complet :
    - Une feuille par catégorie
    - Une feuille Résumé Général
    - Une feuille Produits Individuels
    - Bénéfice masqué si utilisateur est Gerante
    """
    afficher_benefice = request.user.type_utilisateur != 'Gerante'
    lignes = LigneVente.objects.select_related('produit', 'produit__categorie').all()

    # 🔹 Grouper par catégorie
    categories = {}
    for lv in lignes:
        cat = lv.produit.categorie.desgcategorie
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(lv)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    total_global = 0
    total_quantite_global = 0
    total_quantite_retournee_global = 0
    total_benefice_global = 0

    # ================= FEUILLE PAR CATEGORIE =================
    for cat_name, lignes_cat in categories.items():
        ws = wb.create_sheet(title=cat_name[:31])
        headers = ["Produit", "Quantité Vendue", "Quantite Retournée", "Prix Unitaire", "Montant Réduction", "Sous-Total"]
        if afficher_benefice:
            headers.append("Bénéfice")
        headers.append("Total Vente par Produit")

        # En-têtes
        for col_num, header in enumerate(headers, 1):
            cell = ws[f"{get_column_letter(col_num)}1"]
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill("solid", fgColor="4F81BD")
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                 top=Side(style="thin"), bottom=Side(style="thin"))

        # Remplissage lignes
        ligne_excel = 2
        total_categorie = 0
        total_quantite_categorie = 0
        total_quantite_retournee_categorie = 0
        total_benefice_categorie = 0

        for lv in lignes_cat:
            ws[f"A{ligne_excel}"] = lv.produit.desgprod
            ws[f"B{ligne_excel}"] = lv.quantite
            ws[f"C{ligne_excel}"] = lv.quantite_retournee
            ws[f"D{ligne_excel}"] = lv.prix
            ws[f"E{ligne_excel}"] = lv.montant_reduction
            ws[f"F{ligne_excel}"] = lv.sous_total

            for col in ['C','D','E','F']:
                ws[f"{col}{ligne_excel}"].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

            col_benefice = None
            if afficher_benefice:
                col_benefice = 'H'
                ws[f"{col_benefice}{ligne_excel}"] = lv.benefice
                ws[f"{col_benefice}{ligne_excel}"].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

            col_total = 'H' if afficher_benefice else 'G'
            ws[f"{col_total}{ligne_excel}"] = lv.sous_total
            ws[f"{col_total}{ligne_excel}"].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

            total_categorie += lv.sous_total
            total_quantite_categorie += lv.quantite
            total_quantite_retournee_categorie += lv.quantite_retournee
            if afficher_benefice:
                total_benefice_categorie += lv.benefice

            ligne_excel += 1

        # Ligne total par catégorie
        ws[f"A{ligne_excel}"] = f"TOTAL {cat_name}"
        ws[f"A{ligne_excel}"].font = Font(bold=True)
        ws[f"A{ligne_excel}"].alignment = Alignment(horizontal="right")
        ws[f"B{ligne_excel}"] = total_quantite_categorie
        ws[f"C{ligne_excel}"] = total_quantite_retournee_categorie
        ws[f"{col_total}{ligne_excel}"] = total_categorie
        ws[f"{col_total}{ligne_excel}"].font = Font(bold=True, color="FF0000")
        if afficher_benefice:
            ws[f"{col_benefice}{ligne_excel}"] = total_benefice_categorie
            ws[f"{col_benefice}{ligne_excel}"].font = Font(bold=True, color="008000")

        # Ajuster largeur colonnes
        for col in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Totaux globaux
        total_global += total_categorie
        total_quantite_global += total_quantite_categorie
        total_quantite_retournee_global += total_quantite_retournee_categorie
        total_benefice_global += total_benefice_categorie if afficher_benefice else 0

    # ================= FEUILLE PRODUITS INDIVIDUELS =================
    ws_prod = wb.create_sheet(title="Produits Individuels")
    headers_prod = ["Produit", "Catégorie", "Quantité Vendue", "Quantite Retournée", "Prix Unitaire",
                    "Montant Réduction", "Sous-Total"]
    if afficher_benefice:
        headers_prod.append("Bénéfice")
    headers_prod.append("Total Vente par Produit")

    for col_num, header in enumerate(headers_prod, 1):
        cell = ws_prod[f"{get_column_letter(col_num)}1"]
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill("solid", fgColor="4F81BD")

    ligne_excel = 2
    for lv in lignes:
        ws_prod[f"A{ligne_excel}"] = lv.produit.desgprod
        ws_prod[f"B{ligne_excel}"] = lv.produit.categorie.desgcategorie
        ws_prod[f"C{ligne_excel}"] = lv.quantite
        ws_prod[f"D{ligne_excel}"] = lv.quantite_retournee
        ws_prod[f"E{ligne_excel}"] = lv.prix
        ws_prod[f"F{ligne_excel}"] = lv.montant_reduction
        ws_prod[f"G{ligne_excel}"] = lv.sous_total

        for col in ['D','E','F','G']:
            ws_prod[f"{col}{ligne_excel}"].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        col_benefice = None
        if afficher_benefice:
            col_benefice = 'H'
            ws_prod[f"{col_benefice}{ligne_excel}"] = lv.benefice
            ws_prod[f"{col_benefice}{ligne_excel}"].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        col_total = 'I' if afficher_benefice else 'H'
        ws_prod[f"{col_total}{ligne_excel}"] = lv.sous_total
        ws_prod[f"{col_total}{ligne_excel}"].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        ligne_excel += 1

    for col in range(1, len(headers_prod)+1):
        ws_prod.column_dimensions[get_column_letter(col)].width = 20

    # ================= FEUILLE RESUME GENERAL =================
    ws_summary = wb.create_sheet(title="Résumé Général")
    summary_headers = ["Catégorie", "Quantité Totale", "Quantite Retournée", "Total Vente"]
    if afficher_benefice:
        summary_headers.append("Bénéfice Total")

    for col_num, header in enumerate(summary_headers, 1):
        cell = ws_summary[f"{get_column_letter(col_num)}1"]
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill("solid", fgColor="4F81BD")

    ligne_excel = 2
    for cat_name, lignes_cat in categories.items():
        total_cat = sum(lv.sous_total for lv in lignes_cat)
        total_quant_cat = sum(lv.quantite for lv in lignes_cat)
        total_quant_cat_retournee = sum(lv.quantite_retournee for lv in lignes_cat)
        total_benef_cat = sum(lv.benefice for lv in lignes_cat) if afficher_benefice else 0

        ws_summary[f"A{ligne_excel}"] = cat_name
        ws_summary[f"B{ligne_excel}"] = total_quant_cat
        ws_summary[f"C{ligne_excel}"] = total_quant_cat_retournee
        ws_summary[f"D{ligne_excel}"] = total_cat
        if afficher_benefice:
            ws_summary[f"E{ligne_excel}"] = total_benef_cat
        ligne_excel += 1

    # Ligne total général
    ws_summary[f"A{ligne_excel}"] = "TOTAL GENERAL"
    ws_summary[f"A{ligne_excel}"].font = Font(bold=True)
    ws_summary[f"A{ligne_excel}"].alignment = Alignment(horizontal="right")
    ws_summary[f"B{ligne_excel}"] = total_quantite_global
    ws_summary[f"C{ligne_excel}"] = total_quantite_retournee_global
    ws_summary[f"D{ligne_excel}"] = total_global
    if afficher_benefice:
        ws_summary[f"E{ligne_excel}"] = total_benefice_global

    for col in range(1, len(summary_headers)+1):
        ws_summary.column_dimensions[get_column_letter(col)].width = 20

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=ventes_complet.xlsx'
    wb.save(response)
    return response


#================================================================================================
# Fonction pour afficher le formulaire de formulaire d'exportation des données
#================================================================================================
@login_required
def confirmation_exportation_retours_vente(request):
    
    return render(request, 'gestion_produits/exportation/confirmation_exportation_retours_ventes.html')

#================================================================================================
# Fonction pour exporter les retours de vente en Excel
#================================================================================================
@login_required
def export_retours_ventes_excel(request):
    """
    Export des retours de vente :
    - Une feuille par produit
    - Une feuille résumé général
    """

    # 🔹 Récupérer toutes les lignes avec leurs retours
    lignes = LigneVente.objects.select_related('produit', 'vente', 'produit__categorie').all()

    # 🔹 Créer le classeur Excel
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Supprimer la feuille par défaut

    # ====================== FEUILLE DÉTAIL DES RETOURS ======================
    ws_detail = wb.create_sheet(title="Retours Détails")
    headers = ["Code Vente", "Produit", "Catégorie", "Qté Vendue", "Qté Retournée", "Quantité Restante", "Motif", "Date Retour"]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_detail[f"{get_column_letter(col_num)}1"]
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))

    ligne_excel = 2
    for lv in lignes:
        retours = lv.retours.all()
        for ret in retours:
            ws_detail[f"A{ligne_excel}"] = lv.vente.code
            ws_detail[f"B{ligne_excel}"] = lv.produit.desgprod
            ws_detail[f"C{ligne_excel}"] = lv.produit.categorie.desgcategorie
            ws_detail[f"D{ligne_excel}"] = lv.quantite
            ws_detail[f"E{ligne_excel}"] = ret.quantite_retour
            ws_detail[f"F{ligne_excel}"] = lv.quantite - ret.quantite_retour
            ws_detail[f"G{ligne_excel}"] = ret.motif
            ws_detail[f"H{ligne_excel}"] = ret.date_retour.strftime("%d/%m/%Y %H:%M")
            
            for col in ['D','E','F']:
                ws_detail[f"{col}{ligne_excel}"].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

            ligne_excel += 1

    for col in range(1, len(headers)+1):
        ws_detail.column_dimensions[get_column_letter(col)].width = 20

    # ====================== FEUILLE RÉSUMÉ PAR PRODUIT ======================
    ws_summary = wb.create_sheet(title="Résumé Retours")
    summary_headers = ["Produit", "Catégorie", "Total Vendue", "Total Retournée", "Total Restante"]
    
    for col_num, header in enumerate(summary_headers, 1):
        cell = ws_summary[f"{get_column_letter(col_num)}1"]
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill("solid", fgColor="4F81BD")

    ligne_excel = 2
    produits = {}
    for lv in lignes:
        key = lv.produit.id
        if key not in produits:
            produits[key] = {
                "produit": lv.produit,
                "categorie": lv.produit.categorie,
                "total_vendue": 0,
                "total_retournee": 0
            }
        produits[key]["total_vendue"] += lv.quantite
        produits[key]["total_retournee"] += lv.retours.aggregate(total=Sum('quantite_retour'))['total'] or 0

    for p in produits.values():
        ws_summary[f"A{ligne_excel}"] = p["produit"].desgprod
        ws_summary[f"B{ligne_excel}"] = p["categorie"].desgcategorie
        ws_summary[f"C{ligne_excel}"] = p["total_vendue"]
        ws_summary[f"D{ligne_excel}"] = p["total_retournee"]
        ws_summary[f"E{ligne_excel}"] = p["total_vendue"] - p["total_retournee"]

        for col in ['C','D','E']:
            ws_summary[f"{col}{ligne_excel}"].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        ligne_excel += 1

    for col in range(1, len(summary_headers)+1):
        ws_summary.column_dimensions[get_column_letter(col)].width = 20

    # ====================== EXPORT ======================
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=retours_ventes.xlsx'
    wb.save(response)
    return response

#================================================================================================
# Fonction pour afficher le formulaire de formulaire d'exportation des données
#================================================================================================
@login_required
def confirmation_exportation_categorie(request):
    return render(request, 'gestion_produits/exportation/confirmation_exportation_categories.html')

#=============================================================================================
# Fonction pour exporter les données des Catégories Produits
#==============================================================================================
@login_required
def export_categories_excel(request):
    # 1️⃣ Récupération des catégorie (OPTIMISÉ)
    categories = CategorieProduit.objects.all()

    # 2️⃣ Création du fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liste des Catégories Produits"

    # 3️⃣ En-têtes
    headers = [
        "Catégorie",
        "Description",
        "Date de Mise à Jour"
    ]

    for col_num, header in enumerate(headers, 1):
        ws[f"{get_column_letter(col_num)}1"] = header

    # 4️⃣ Données
    ligne = 2
    for elems in categories:
        ws[f"A{ligne}"] = elems.desgcategorie
        ws[f"B{ligne}"] = elems.description
        ws[f"C{ligne}"] = elems.date_maj.strftime("%d/%m/%Y %H:%M") if elems.date_maj else ""
        ligne += 1

    # 5️⃣ Ajuster largeur colonnes
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25

    # 6️⃣ Téléchargement
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=liste_categories_produits.xlsx'
    wb.save(response)
    return response

#================================================================================================
# Fonction pour afficher le formulaire de formulaire d'exportation des données
#================================================================================================
@login_required
def confirmation_exportation_produits(request):
    
    return render(request, 'gestion_produits/exportation/confirmation_exportation_produits.html')

#=============================================================================================
# Fonction pour exporter les données des ventes
#==============================================================================================
@login_required
def export_produits_excel(request):
    # 1️⃣ Récupération des produits + catégorie (OPTIMISÉ)
    produits = Produits.objects.select_related('categorie').all()

    # 2️⃣ Création du fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liste des Produits"

    # 3️⃣ En-têtes
    headers = [
        "Catégorie",
        "Référence Produit",
        "Désignation",
        "Prix Unitaire",
        "Quantité en Stock",
        "Seuil",
        "Date de Mise à Jour"
    ]

    for col_num, header in enumerate(headers, 1):
        ws[f"{get_column_letter(col_num)}1"] = header

    # 4️⃣ Données
    ligne = 2
    for produit in produits:
        ws[f"A{ligne}"] = produit.categorie.desgcategorie if produit.categorie else ""
        ws[f"B{ligne}"] = produit.refprod
        ws[f"C{ligne}"] = produit.desgprod
        ws[f"D{ligne}"] = produit.pu
        ws[f"E{ligne}"] = produit.qtestock
        ws[f"F{ligne}"] = produit.seuil
        ws[f"G{ligne}"] = produit.date_maj.strftime("%d/%m/%Y %H:%M") if produit.date_maj else ""
        ligne += 1

    # 5️⃣ Ajuster largeur colonnes
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25

    # 6️⃣ Téléchargement
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=liste_produits.xlsx'
    wb.save(response)
    return response

#================================================================================================
# Fonction pour afficher le formulaire de formulaire d'exportation des données
#================================================================================================
@login_required
def confirmation_exportation_stocks_produits(request):
    
    return render(request, 'gestion_produits/exportation/confirmation_exportation_stocksproduits.html')
#=============================================================================================
# Fonction pour exporter les données des stocks avec résumé par catégorie
#==============================================================================================
@login_required
def export_stocks_excel_resume(request):
    # 1. Vérifier si on doit afficher le prix en gros
    show_prix_en_gros = request.user.type_utilisateur != 'Gerante'

    # 2. Récupérer tous les stocks avec les relations nécessaires
    stocks = StockProduit.objects.select_related('produit', 'produit__categorie').all()

    # 3. Créer le fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stocks Produits"

    # 4. Définir les en-têtes
    headers = ["Produit", "Catégorie", "Prix de Vente Unitaire", "Stock actuel"]
    if show_prix_en_gros:
        headers.append("Prix Unitaire en Gros")
    headers.append("Valeur Stock")

    for col_num, header in enumerate(headers, 1):
        ws[f"{get_column_letter(col_num)}1"] = header

    # 5. Insérer les lignes produits
    ligne = 2
    for sp in stocks:
        data = [
            sp.produit.desgprod,
            sp.produit.categorie.desgcategorie,
            sp.produit.pu,       # <-- Prix de vente unitaire
            sp.qtestock          # <-- Stock actuel
        ]
        if show_prix_en_gros:
            data.append(sp.produit.prix_en_gros)
        data.append(sp.qtestock * sp.produit.pu)  # Valeur stock

        for col_num, value in enumerate(data, 1):
            ws[f"{get_column_letter(col_num)}{ligne}"] = value
        ligne += 1

    # 6. Ajouter un résumé par catégorie
    ws_summary = wb.create_sheet(title="Résumé par Catégorie")

    summary_headers = ["Catégorie", "Nombre de Produits", "Stock Total", "Prix de Vente Unitaire", "Valeur Stock Totale"]
    if show_prix_en_gros:
        summary_headers.insert(3, "Prix Unitaire Moyen en Gros")

    for col_num, header in enumerate(summary_headers, 1):
        ws_summary[f"{get_column_letter(col_num)}1"] = header

    # Calcul des totaux par catégorie
    categories = stocks.values(
        'produit__categorie__desgcategorie'
    ).annotate(
        nombre_produits=Count('produit', distinct=True),
        stock_total=Sum('qtestock'),
        valeur_stock_totale=Sum(F('qtestock') * F('produit__pu')),
        prix_en_gros_moyen=Sum('produit__prix_en_gros') / Count('produit')
    ).order_by('produit__categorie__desgcategorie')

    ligne = 2
    for cat in categories:
        data = [
            cat['produit__categorie__desgcategorie'],
            cat['nombre_produits'],
            cat['stock_total'],
            # Prix de vente unitaire moyen par catégorie
            round(cat['valeur_stock_totale'] / cat['stock_total'], 2) if cat['stock_total'] else 0
        ]
        if show_prix_en_gros:
            data.insert(3, round(cat['prix_en_gros_moyen'], 2))
        data.append(cat['valeur_stock_totale'])

        for col_num, value in enumerate(data, 1):
            ws_summary[f"{get_column_letter(col_num)}{ligne}"] = value
        ligne += 1

    # 7. Ajuster largeur des colonnes pour les deux feuilles
    for sheet in [ws, ws_summary]:
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 25

    # 8. Retourner le fichier Excel
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=stocks_produits_resume.xlsx'
    wb.save(response)
    return response

#================================================================================================
# Fonction pour afficher le formulaire de formulaire d'exportation des données
#================================================================================================
@login_required
def confirmation_exportation_commande(request):
    
    return render(request, 'gestion_produits/exportation/confirmation_exportation_commandes.html')

@login_required
def export_commandes_excel(request):
    # 1. Récupérer les commandes avec les produits et catégories
    commandes = Commandes.objects.select_related(
        'produits',
        'produits__categorie'
    ).order_by('-datecmd')

    # 2. Créer le fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liste des Commandes"

    # 3. En-têtes
    headers = [
        "N° Commande",
        "Date Commande",
        "Produit",
        "Catégorie",
        "Quantité Commandée",
        "Fournisseur",
        "Téléphone Fournisseur"
    ]

    for col, header in enumerate(headers, 1):
        ws[f"{get_column_letter(col)}1"] = header

    # 4. Remplir les lignes
    ligne = 2
    for cmd in commandes:
        ws[f"A{ligne}"] = cmd.numcmd
        ws[f"B{ligne}"] = cmd.datecmd.strftime("%d/%m/%Y")
        ws[f"C{ligne}"] = cmd.produits.desgprod
        ws[f"D{ligne}"] = (
            cmd.produits.categorie.desgcategorie
            if cmd.produits.categorie else ""
        )
        ws[f"E{ligne}"] = cmd.qtecmd
        ws[f"F{ligne}"] = cmd.nom_complet_fournisseur or ""
        ws[f"G{ligne}"] = cmd.telephone_fournisseur or ""
        ligne += 1

    # 5. Ajuster largeur colonnes
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25

    # 6. Réponse HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = (
        f'attachment; filename=commandes_{timezone.now().date()}.xlsx'
    )

    wb.save(response)
    return response

#================================================================================================
# Fonction pour afficher le formulaire de formulaire d'exportation des données
#================================================================================================
@login_required
def confirmation_exportation_livraison(request):
    
    return render(request, 'gestion_produits/exportation/confirmation_exportation_livraisons.html')

@login_required
def export_livraisons_excel(request):
    # Récupérer toutes les livraisons avec les produits
    livraisons = LivraisonsProduits.objects.select_related('produits').all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Livraisons"

    # En-têtes incluant les infos commande
    headers = [
        "Produit", "Quantité Livrée", "Date Livraison", "Statut",
        "Numéro Commande", "Quantité Commandée",
        "Fournisseur", "Téléphone Fournisseur", "Adresse Fournisseur"
    ]

    for col, header in enumerate(headers, 1):
        ws[f"{get_column_letter(col)}1"] = header

    ligne = 2
    for l in livraisons:
        # Tenter de récupérer la commande associée au produit et à la date de livraison
        commande = Commandes.objects.filter(
            produits=l.produits
        ).order_by('-datecmd').first()  # On prend la dernière commande pour ce produit

        ws[f"A{ligne}"] = l.produits.desgprod
        ws[f"B{ligne}"] = l.qtelivrer
        ws[f"C{ligne}"] = l.datelivrer.strftime("%d/%m/%Y")
        ws[f"D{ligne}"] = l.statuts

        if commande:
            ws[f"E{ligne}"] = commande.numcmd
            ws[f"F{ligne}"] = commande.qtecmd
            ws[f"G{ligne}"] = commande.nom_complet_fournisseur
            ws[f"H{ligne}"] = commande.telephone_fournisseur
            ws[f"I{ligne}"] = commande.adresse_fournisseur
        else:
            ws[f"E{ligne}"] = ""
            ws[f"F{ligne}"] = ""
            ws[f"G{ligne}"] = ""
            ws[f"H{ligne}"] = ""
            ws[f"I{ligne}"] = ""

        ligne += 1

    # Ajuster la largeur des colonnes
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=livraisons.xlsx"
    wb.save(response)
    return response

#================================================================================================
