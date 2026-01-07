from io import BytesIO
from urllib import request
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Q, F
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.urls import reverse
from django.db import DatabaseError, IntegrityError
from gestion_audit.views import enregistrer_audit
from gestion_notifications.models import Notification
from .models import *
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.db import DatabaseError, IntegrityError
import qrcode
import base64
from django.core.mail import send_mail, EmailMessage
from .utils import pagination_liste
from django.conf import settings
from django.db import transaction, IntegrityError, DatabaseError
from django.core.mail import send_mail
from datetime import datetime
from decimal import Decimal, InvalidOperation
from django.shortcuts import get_object_or_404, render, redirect
from openpyxl.utils import get_column_letter
from django.db import transaction
import openpyxl

import logging
logger = logging.getLogger(__name__)

from collections import defaultdict 

@login_required(login_url='gestionUtilisateur:connexion_utilisateur') #Empecher tant que l'utilisateur n'est pas connecté
def nouvelle_saisie(request, *args, **kwargs):
    template_name = "gestion_entreprise/listes_entreprise.html"
    # Vérification s'il existe déjà un établissement
    if Entreprise.objects.exists():
        messages.warning(request, "⚠️ Une entreprise de vente existe déjà. Vous ne pouvez pas en ajouter un autre.")
        return redirect('liste_entreprise') 

    if request.method == 'POST':
        nom_entrepriese = request.POST.get('nom_entrepriese')
        data = {
            'nom_entrepriese': nom_entrepriese,
            'email': request.POST.get('email'),
            'adresse': request.POST.get('adresse'),
            'contact1': request.POST.get('contact1'), 
            'contact2': request.POST.get('contact2'), 
            'logo': request.FILES.get('logo'),
        }

        try:
            Entreprise.objects.create(**data)
            messages.success(request, "L'entreprise a été ajouté avec succès.")
            return redirect('liste_entreprise')
        except Exception as ex:
            messages.error(request, f"❌ Erreur d'insertion de l'entreprise : {str(ex)}")
            return render(request, template_name, {'data': data})

    return render(request, template_name)

#==================================================================================================================
#Liste Etablissement Scolaire
#==================================================================================================================
def liste_entreprise(request, *args, **kwargs):
    liste_entreprise = Entreprise.objects.all().order_by('id')
    context = {
        'liste_entreprise' : liste_entreprise
    }
    return render(request, 'gest_entreprise/listes_entreprise.html', context)

#==================================================================================================================
#Fonction pour supprimer un Entreprise
#==================================================================================================================
@login_required
def supprimer_entreprise(request):
    try:
        identifiant = request.POST.get('id_supprimer')
        etablissement = get_object_or_404(Entreprise, id=identifiant)
        etablissement.delete()
        messages.success(request, "Suppression effectuée avec succès !")
        return redirect('liste_etablissement')
    except Exception as ex:
        messages.error(request, f"Erreur de Suppression {ex}")
    return render(request, "gest_entreprise/listes_entreprise.html")
#==================================================================================================================
#Fonction Pour Modifier
#==================================================================================================================
@login_required
def modifier_entreprise(request):
    try:
            id_modif = request.POST.get('id_modif')
            entreprise = get_object_or_404(Entreprise, id=id_modif)
            
            entreprise.nom_entrepriese = request.POST.get("nom_entrepriese")
            entreprise.adresse = request.POST.get("adresse")
            entreprise.email = request.POST.get("email")
            entreprise.contact1 = request.POST.get("contact1")
            entreprise.contact2 = request.POST.get("contact2")
            
            if request.FILES.get("logo"):
                entreprise.logo = request.FILES.get("logo")
            entreprise.save()
            messages.success(request, "Modification effectuée avec succès ! ")
            return redirect('liste_entreprise')
    except Exception as ex:
            messages.warning(request, f"Erreur de Modiication des Informations {ex}")
    return redirect('liste_entreprise')


# =================================================================================================
# Liste des Dépenses
# =================================================================================================
@login_required
def liste_depense(request):
    try:
        depenses_list = Depenses.objects.all().order_by("-date_operation")
        total_depenses = depenses_list.count()
        
        paginator = Paginator(depenses_list, 10)
        page = request.GET.get("page")
        liste_depenses = paginator.get_page(page)

        context = {
            "liste_depenses": liste_depenses,
            "total_depenses": total_depenses,
        }
        return render(request, "gest_entreprise/depenses/listes_depenses.html", context)
    except DatabaseError as db_err:
        messages.error(request, f"Erreur de base de données : {str(db_err)}")
    except Exception as e:
        messages.error(request, f"Erreur inattendue : {str(e)}")
    return redirect("liste_depense")

# =================================================================================================
# Fonction pour filtrer la Liste des Dépenses
# =================================================================================================
@login_required
def filtrer_listes_depenses(request):
    """
    Filtre les dépenses de l'établissement selon la période choisie
    et calcule la somme totale des montants filtrés.
    """
    try:
        # --- Récupération de toutes les dépenses ---
        listes_depenses = Depenses.objects.all()

        # --- Récupération des dates dans la requête ---
        date_debut = request.GET.get("date_debut")
        date_fin = request.GET.get("date_fin")

        # --- Application du filtre par date ---
        if date_debut and date_fin:
            listes_depenses = listes_depenses.filter(
                date_operation__range=(date_debut, date_fin)
            )

        # --- Pagination ---
        listes_depenses_pagine = pagination_liste(request, listes_depenses)

        # --- Calculs statistiques ---
        total_depenses = listes_depenses.count()  # nombre total des dépenses filtrées

    except Exception as ex:
        messages.warning(request, f"⚠️ Erreur lors du filtrage des données : {str(ex)}")
        listes_depenses_pagine = []
        total_depenses = 0
        date_debut = None
        date_fin = None

    # --- Contexte pour le template ---
    context = {
        "date_debut": date_debut,
        "date_fin": date_fin,
        "listes_depenses_pagine": listes_depenses_pagine,
        "total_depenses": total_depenses,
    }

    return render(request, "gest_entreprise/depenses/listes_depenses.html", context)

# =================================================================================================
# Ajouter une nouvelle dépense
# =================================================================================================

@login_required
def nouvelle_depense(request):
    if request.method != "POST":
        messages.warning(request, "⚠️ Méthode non autorisée.")
        return redirect("liste_depense")

    try:
        designation = request.POST.get("designation", "").strip()
        destine = request.POST.get("destine_a", "").strip()
        montant = request.POST.get("montant", "").strip()

        if not all([designation, destine, montant]):
            messages.error(request, "⚠️ Tous les champs doivent être remplis.")
            return redirect("liste_depense")

        montant_decimal = Decimal(montant)

        depense = Depenses.objects.create(
            designation=designation,
            destine_a=destine,
            montant=montant_decimal,
            utilisateur=request.user
        )

        # Envoi email
        try:
            sujet = "🧾 Nouvelle dépense enregistrée"
            message = (
                f"Une nouvelle dépense a été ajoutée par {request.user.get_full_name()}.\n\n"
                f"Détails :\n"
                f"- Désignation : {designation}\n"
                f"- Destinée à : {destine}\n"
                f"- Montant : {montant} GNF\n"
            )
            destinataires = [settings.ADMIN_EMAIL] if hasattr(settings, "ADMIN_EMAIL") else ["admin@etablissement.com"]

            send_mail(sujet, message, settings.DEFAULT_FROM_EMAIL, destinataires, fail_silently=False)
        except Exception as email_error:
            messages.warning(request, f"📧 Dépense enregistrée mais erreur d’e-mail : {email_error}")

        messages.success(request, f"✅ Dépense enregistrée avec succès GNF.")

    except (IntegrityError, DatabaseError) as db_err:
        messages.error(request, f"❌ Erreur de base de données : {db_err}")
    except Exception as e:
        messages.error(request, f"❌ Erreur inattendue : {e}")

    return redirect("liste_depense")

#=============================================================================================
# Fonction pour gérer les réçu des Dépenses
#=============================================================================================

@login_required
def recu_depense(request, depense_id):
    """
    Affiche le reçu détaillé d'une dépense avec QR code.
    """
    try:
        depense = get_object_or_404(Depenses, id=depense_id)
        nom_entreprise = Entreprise.objects.first()
        today = timezone.now()

        # --- QR code ---
        qr_text = (
            f"Dépense: {depense.designation}\n"
            f"Montant: {depense.montant:,.0f} GNF\n"
            f"Destiné à: {depense.destine_a}\n"
            f"Utilisateur: {depense.utilisateur}\n"
            f"Date: {depense.date_operation.strftime('%Y-%m-%d') if depense.date_operation else '-'}"
        )
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=10,
            border=4
        )
        qr.add_data(qr_text)
        qr.make(fit=True)
        buffer = BytesIO()
        qr.make_image(fill_color="black", back_color="white").save(buffer, format="PNG")
        qr_code_base64 = base64.b64encode(buffer.getvalue()).decode()

        context = {
            "depense": depense,
            "nom_entreprise": nom_entreprise,
            "today": today,
            "qr_code_base64": qr_code_base64,
            "montant": depense.montant,
        }

        return render(request, "gest_entreprise/depenses/recu_depense/recu_depense.html", context)

    except Exception as ex:
        messages.error(request, f"❌ Erreur lors de la récupération du reçu de dépense : {str(ex)}")
        return redirect("liste_depenses")

#=============================================================================================
# Fonction pour gérer les réçu de dépenses global dans un interval
# #=============================================================================================
@login_required
def recu_depense_global_interval(request):
    """
    Affiche un reçu global pour toutes les dépenses entre deux dates passées en GET.
    """
    try:
        # --- Récupération des dates dans la requête GET ---
        date_debut = request.GET.get("date_debut")
        date_fin = request.GET.get("date_fin")

        if not date_debut or not date_fin:
            messages.error(request, "⚠️ Veuillez fournir une date de début et une date de fin.")
            return redirect("liste_depenses")  # <-- Vérifie que ce nom existe dans urls.py

        # --- Conversion des dates ---
        date_debut_obj = datetime.strptime(date_debut, "%Y-%m-%d").date()
        date_fin_obj = datetime.strptime(date_fin, "%Y-%m-%d").date()

        if date_debut_obj > date_fin_obj:
            messages.error(request, "⚠️ La date de début doit être antérieure à la date de fin.")
            return redirect("liste_depense")

        # --- Récupérer les dépenses dans l'intervalle ---
        depenses = Depenses.objects.filter(
            date_operation__gte = date_debut_obj,
            date_operation__lte = date_fin_obj
        ).order_by("date_operation")

        if not depenses.exists():
            messages.warning(request, "⚠️ Aucune dépense trouvée dans cet intervalle.")
            return redirect("liste_depenses")

        # --- Calcul du total des dépenses ---
        total_depenses = sum((d.montant or Decimal("0.00")) for d in depenses)

        # --- Génération QR code global ---
        qr_text = f"Dépenses totales entre {date_debut_obj.strftime('%d/%m/%Y')} et {date_fin_obj.strftime('%d/%m/%Y')}\n"
        for d in depenses:
            qr_text += f"- {d.designation}: {d.montant:,.0f} GNF\n"
        qr_text += f"Total: {total_depenses:,.0f} GNF"

        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_text)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buffer = BytesIO()
        img.save(buffer, format="PNG")
        qr_code_base64 = base64.b64encode(buffer.getvalue()).decode()

        # --- Profil établissement et date du jour ---
        nom_entreprise = Entreprise.objects.first()
        today = datetime.now()

    except Exception as ex:
        messages.error(request, f"❌ Erreur lors de la récupération des dépenses : {str(ex)}")
        return redirect("liste_depenses")

    context = {
        "depenses": depenses,
        "total_depenses": total_depenses,
        "info_entreprise": nom_entreprise,
        "today": today,
        "qr_code_base64": qr_code_base64,
        "date_debut": date_debut_obj,
        "date_fin": date_fin_obj,
    }

    return render(
        request,
        "gest_entreprise/depenses/recu_depense/recu_depenses_global_interval.html",
        context
    )

# =================================================================================================
# Modifier une dépense existante
# =================================================================================================
@login_required
def modifier_depense(request):
    """
    Modifie une dépense existante,
    enregistre un audit
    et notifie l’administration par email.
    """

    if request.method != "POST":
        return redirect("liste_depense")

    try:
        # 🔹 Récupération de la dépense
        id_depense = request.POST.get("id_modif")
        depense = get_object_or_404(Depenses, pk=id_depense)

        # 🔹 Anciennes valeurs (audit)
        ancienne_valeur = {
            "designation": depense.designation,
            "montant": depense.montant,
            "destine_a": depense.destine_a,
            "utilisateur": depense.utilisateur.username if depense.utilisateur else "Inconnu",
        }

        # 🔹 Nouvelles valeurs
        designation = request.POST.get("modif_designation", "").strip()
        destine = request.POST.get("modif_destine", "").strip()
        montant_brut = request.POST.get("modif_montant", "").strip()
        print(f"montant_brut: {montant_brut}")

        # ❌ Validation champs
        if not designation or not destine or not montant_brut:
            messages.error(request, "⚠️ Tous les champs sont obligatoires.")
            return redirect("liste_depense")

        # 🔒 Validation montant
        montant_brut = request.POST.get("modif_montant", "").strip()

        try:
                # Remplacer la virgule par un point (format FR → format Python)
            montant_nettoye = montant_brut.replace(" ", "").replace(",", ".")

            montant = float(montant_nettoye)

            if montant <= 0:
                raise ValueError

        except ValueError:
            messages.error(request, "⚠️ Le montant doit être un nombre positif valide.")
            return redirect("liste_depense")

        # ================= TRANSACTION =================
        with transaction.atomic():
            depense.designation = designation
            depense.destine_a = destine
            depense.montant = montant
            depense.utilisateur = request.user   # ✅ INSTANCE User
            depense.save()

            # 🔹 Audit
            enregistrer_audit(
                utilisateur=request.user,
                action="Modification",
                table="Depenses",
                ancienne_valeur=ancienne_valeur,
                nouvelle_valeur={
                    "designation": designation,
                    "montant": montant,
                    "destine_a": destine,
                    "utilisateur": request.user.username,
                },
            )

        # ================= EMAIL =================
        try:
            destinataires = [settings.ADMIN_EMAIL] if hasattr(settings, "ADMIN_EMAIL") else []

            if destinataires:
                send_mail(
                    "✏️ Modification d’une dépense",
                    f"""Une dépense a été modifiée.

                Utilisateur : {request.user.username}

                Anciennes valeurs :
                - Désignation : {ancienne_valeur['designation']}
                - Montant : {ancienne_valeur['montant']} GNF

                Nouvelles valeurs :
                - Désignation : {designation}
                - Montant : {montant} GNF
                - Destinée à : {destine}
                """,
                    settings.DEFAULT_FROM_EMAIL,
                    destinataires,
                    fail_silently=False,
                )

        except Exception:
            messages.warning(request, "📧 Dépense modifiée mais e-mail non envoyé.")

        messages.success(request, "✅ Dépense modifiée avec succès.")

    except DatabaseError:
        messages.error(request, "⚠️ Erreur de base de données.")
    except Exception as e:
        messages.error(request, f"⚠️ Erreur inattendue : {e}")

    return redirect("liste_depense")


# =================================================================================================
# Supprimer une dépense
# =================================================================================================
@login_required
def supprimer_depense(request):
    if request.method != "POST":
        return redirect("liste_depense")

    try:
        id_depense = request.POST.get("id_supprimer")
        depense = get_object_or_404(Depenses, pk=id_depense)

        # 🔒 Sécurité : seul le créateur peut supprimer
        if hasattr(depense, "utilisateur") and depense.utilisateur:
            if depense.utilisateur.id != request.user.id:
                messages.warning(
                    request,
                    "❌ Vous ne pouvez pas supprimer cette dépense : "
                    f"elle a été enregistrée par {depense.utilisateur.get_full_name()}."
                )
                return redirect("liste_depense")

        # 🔍 Sauvegarde pour audit + email
        ancienne_valeur = {
            "designation": depense.designation,
            "montant": depense.montant,
            "destine_a": depense.destine_a,
            "utilisateur": depense.utilisateur.get_full_name() if depense.utilisateur else "Inconnu",
            "date": depense.date_depense.strftime("%d/%m/%Y %H:%M")
            if hasattr(depense, "date_depense") else "-"
        }

        with transaction.atomic():

            # 🧾 Audit
            enregistrer_audit(
                utilisateur=request.user,
                action="Suppression",
                table="Depenses",
                ancienne_valeur=ancienne_valeur,
                nouvelle_valeur=None
            )

            # ❌ Suppression réelle
            depense.delete()

        # 📧 EMAIL DE NOTIFICATION
        try:
            email_body = (
                f"SUPPRESSION D’UNE DÉPENSE\n\n"
                f"Désignation : {ancienne_valeur['designation']}\n"
                f"Montant : {ancienne_valeur['montant']} GNF\n"
                f"Destinée à : {ancienne_valeur['destine_a']}\n"
                f"Enregistrée par : {ancienne_valeur['utilisateur']}\n"
                f"Supprimée par : {request.user.get_full_name()}\n"
                f"Date : {timezone.now().strftime('%d/%m/%Y %H:%M')}\n"
            )

            EmailMessage(
                subject="Suppression d’une dépense",
                body=email_body,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[settings.ADMIN_EMAIL]
            ).send(fail_silently=False)
            
                    # 6️⃣ Notification interne
            Notification.objects.create(
                destinataire=request.user,
                destinataire_email = settings.ADMIN_EMAIL,
                titre="🗑 Suppression d'une Dépense",
                message=(
                    f"Désignation {depense.designation} \n"
                    f"Supprimée par : {request.user.get_full_name()}"
                )
            )

        except Exception as e:
            logger.warning(f"Email suppression dépense non envoyé : {e}")

        messages.success(request, "✅ Dépense supprimée avec succès.")

    except Depenses.DoesNotExist:
        messages.error(request, "❌ La dépense spécifiée n’existe pas.")
    except Exception as e:
        messages.error(request, f"❌ Erreur inattendue : {e}")

    return redirect("liste_depense")

# =================================================================================================
@login_required
def modal_exportation_excel(request):
    return render(request, 'gest_entreprise/depenses/exportation/exportation_donnees_excel.html')


#=============================================================================================
# Fonction pour exporter les données des élèves vers Excel
#==============================================================================================
@login_required(login_url='gestionUtilisateur:connexion_utilisateur') #Empecher tant que l'utilisateur n'est pas connecté
def export_depenses_excel(request):

    #Exporte la liste des depenses au format Excel (.xlsx)

    # 1. Récupérer tous les Eleves
    depense = Depenses.objects.all()

    # 2. Créer un fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liste des Dépenses"

    # 3. Ajouter les en-têtes
    
    ws['A1'] = '#'
    ws['B1'] = 'Date Opération'
    ws['C1'] = 'Montant'
    ws['D1'] = 'Designation'
    ws['E1'] = 'Destiné A'


    # 4. Insérer les données ligne par ligne
    ligne = 2
    for elems in depense :
        ws[f'A{ligne}'] = elems.id
        ws[f'B{ligne}'] = elems.date_operation
        ws[f'C{ligne}'] = elems.montant
        ws[f'D{ligne}'] = elems.designation
        ws[f'E{ligne}'] = elems.destine_a
        ligne += 1

    # 5. Ajuster la largeur des colonnes
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 30

    # 6. Retourner le fichier Excel en téléchargement
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename=gestion_depenses.xlsx'
    wb.save(response)

    return response

#==============================================================================================
# Fonction pour le choix de la date pour l'impression des dépenses
#==============================================================================================
@login_required
def choix_listes_impression_depenses(request):
    return render(request, "gest_entreprise/depenses/impression_listes/choix_impression_depenses.html")

#==============================================================================================
@login_required
def liste_depenses_impression(request):
    date_debut_str = request.POST.get('date_debut')
    date_fin_str = request.POST.get('date_fin')

    if not date_debut_str or not date_fin_str:
        messages.error(request, "⚠️ Veuillez renseigner les deux dates.")
        return render(request, "gest_entreprise/depenses/impression_listes/choix_impression_depenses.html")

    date_debut = datetime.strptime(date_debut_str, "%Y-%m-%d").date()
    date_fin = datetime.strptime(date_fin_str, "%Y-%m-%d").date()

    depenses = Depenses.objects.filter(date_operation__range=(date_debut, date_fin)).order_by('date_operation')

    # Grouper par date
    depenses_par_date_dict = defaultdict(list)
    total_general = 0
    for dep in depenses:
        depenses_par_date_dict[dep.date_operation].append(dep)
        total_general += dep.montant

    # Créer une liste pour le template
    depenses_par_date = [
        {'date': date, 'depenses': items, 'total': sum(d.montant for d in items)}
        for date, items in depenses_par_date_dict.items()
    ]

    context = {
        'depenses_par_date': depenses_par_date,
        'total_general': total_general,
        'nom_entreprise': Entreprise.objects.first(),
        'date_debut': date_debut,
        'date_fin': date_fin,
        'today': timezone.now(),
    }

    return render(request, 'gest_entreprise/depenses/impression_listes/apercue_avant_impression_listes_depenses.html', context)
